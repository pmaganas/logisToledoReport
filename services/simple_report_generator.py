import openpyxl
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from services.sesame_api import SesameAPI

class SimpleReportGenerator:
    def __init__(self):
        self.sesame_api = SesameAPI()
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)

    def generate_simple_report(self, from_date: str = None, to_date: str = None, 
                             employee_id: str = None, office_id: str = None, 
                             department_id: str = None, report_type: str = "by_employee") -> Optional[bytes]:
        """Generate a simplified XLSX report with reduced API calls"""
        
        try:
            self.logger.info(f"Starting simple report generation - Type: {report_type}")
            
            # Get employees data
            if employee_id:
                employee_response = self.sesame_api.get_employee_details(employee_id)
                employees = [employee_response.get('data')] if employee_response else []
            else:
                # Get all employees data with complete pagination
                all_employees = self.sesame_api.get_all_employees_data()
                
                employees = []
                for employee in all_employees:
                    include_employee = True
                    
                    # Filter by office_id
                    if office_id:
                        employee_office_id = employee.get('office', {}).get('id') if employee.get('office') else None
                        if employee_office_id != office_id:
                            include_employee = False
                    
                    # Filter by department_id
                    if department_id and include_employee:
                        employee_department_id = employee.get('department', {}).get('id') if employee.get('department') else None
                        if employee_department_id != department_id:
                            include_employee = False
                    
                    if include_employee:
                        employees.append(employee)
            
            if not employees:
                self.logger.warning("No employees found")
                return self._create_empty_report()
            
            self.logger.info(f"Processing {len(employees)} employees")
            
            # Get check types for activity name resolution
            check_types_data = self.sesame_api.get_all_check_types_data()
            check_types_map = {}
            if check_types_data:
                for check_type in check_types_data:
                    check_types_map[check_type.get('id')] = check_type.get('name', 'Actividad no especificada')
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Simplificado"
            
            # Headers
            headers = ["Empleado", "Tipo ID", "Número ID", "Fecha", "Actividad", "Grupo", "Entrada", "Salida", "Tiempo Registrado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            current_row = 2
            
            # Process based on report type
            if report_type == "by_employee":
                current_row = self._generate_by_employee_report(ws, employees, from_date, to_date, check_types_map, current_row)
            elif report_type == "by_type":
                current_row = self._generate_by_type_report(ws, employees, from_date, to_date, check_types_map, current_row)
            elif report_type == "by_group":
                current_row = self._generate_by_group_report(ws, employees, from_date, to_date, check_types_map, current_row)
            else:
                # Default to by_employee
                current_row = self._generate_by_employee_report(ws, employees, from_date, to_date, check_types_map, current_row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            self.logger.info("Simple report generated successfully")
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error generating simple report: {str(e)}")
            return self._create_error_report(str(e))

    def _generate_by_employee_report(self, ws, employees, from_date, to_date, check_types_map, current_row):
        """Generate report grouped by employee with daily totals"""
        for i, employee in enumerate(employees, 1):
            emp_id = employee.get('id')
            emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
            identity_type, identity_number = self._get_employee_identification(employee)
            
            self.logger.info(f"Processing employee {i}/{len(employees)}: {emp_name}")
            
            try:
                # Get all time tracking data
                time_entries = self.sesame_api.get_all_time_tracking_data(
                    employee_id=emp_id,
                    from_date=from_date,
                    to_date=to_date
                )
                
                if time_entries:
                    # Get break entries for redistribution
                    break_entries = self.sesame_api.get_all_breaks_data(
                        employee_id=emp_id,
                        from_date=from_date,
                        to_date=to_date
                    )
                    
                    # Process break time redistribution
                    processed_entries = self._process_break_redistribution(time_entries, break_entries)
                    
                    # Group entries by date
                    entries_by_date = {}
                    for entry in processed_entries:
                        work_in = entry.get('workEntryIn', {})
                        if work_in and work_in.get('date'):
                            try:
                                parsed_date = self._parse_datetime(work_in.get('date'))
                                if parsed_date:
                                    date_key = parsed_date.strftime('%Y-%m-%d')
                                    if date_key not in entries_by_date:
                                        entries_by_date[date_key] = []
                                    entries_by_date[date_key].append(entry)
                            except Exception:
                                continue
                    
                    # Process each date in chronological order
                    for date_key in sorted(entries_by_date.keys()):
                        date_entries = entries_by_date[date_key]
                        daily_total_seconds = 0
                        activity_types = {}
                        
                        # Add individual entries for this date
                        for entry in date_entries:
                            work_in = entry.get('workEntryIn', {})
                            work_out = entry.get('workEntryOut', {})
                            
                            start_time = self._parse_datetime(work_in.get('date'))
                            end_time = self._parse_datetime(work_out.get('date'))
                            
                            # Calculate duration
                            duration_str = "00:00:00"
                            entry_seconds = 0
                            if start_time and end_time:
                                duration = end_time - start_time
                                entry_seconds = duration.total_seconds()
                                daily_total_seconds += entry_seconds
                                hours = int(entry_seconds // 3600)
                                minutes = int((entry_seconds % 3600) // 60)
                                seconds = int(entry_seconds % 60)
                                duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                            
                            # Get activity name
                            activity_name = "Trabajo"
                            work_check_type_id = entry.get('workCheckTypeId')
                            if work_check_type_id and work_check_type_id in check_types_map:
                                activity_name = check_types_map[work_check_type_id]
                            elif entry.get('workEntryType'):
                                activity_name = entry.get('workEntryType')
                            
                            # Track activity types for totals
                            if activity_name not in activity_types:
                                activity_types[activity_name] = 0
                            activity_types[activity_name] += entry_seconds
                            
                            # Add row for this entry
                            ws.cell(row=current_row, column=1, value=emp_name)
                            ws.cell(row=current_row, column=2, value=identity_type)
                            ws.cell(row=current_row, column=3, value=identity_number)
                            ws.cell(row=current_row, column=4, value=date_key)
                            ws.cell(row=current_row, column=5, value=activity_name)
                            ws.cell(row=current_row, column=6, value="")  # Group column empty
                            ws.cell(row=current_row, column=7, value=start_time.strftime("%H:%M:%S") if start_time else "N/A")
                            ws.cell(row=current_row, column=8, value=end_time.strftime("%H:%M:%S") if end_time else "N/A")
                            ws.cell(row=current_row, column=9, value=duration_str)
                            current_row += 1
                        
                        # Add total row for this date
                        activity_summary = ", ".join([f"{activity}: {int(seconds//3600):02d}:{int((seconds%3600)//60):02d}:{int(seconds%60):02d}" 
                                                     for activity, seconds in activity_types.items()])
                        
                        daily_total_hours = int(daily_total_seconds // 3600)
                        daily_total_minutes = int((daily_total_seconds % 3600) // 60)
                        daily_total_secs = int(daily_total_seconds % 60)
                        daily_total_str = f"{daily_total_hours:02d}:{daily_total_minutes:02d}:{daily_total_secs:02d}"
                        
                        # Add TOTAL row
                        ws.cell(row=current_row, column=1, value="TOTAL")
                        ws.cell(row=current_row, column=2, value="")
                        ws.cell(row=current_row, column=3, value="")
                        ws.cell(row=current_row, column=4, value="")
                        ws.cell(row=current_row, column=5, value=activity_summary)
                        ws.cell(row=current_row, column=6, value="")
                        ws.cell(row=current_row, column=7, value="")
                        ws.cell(row=current_row, column=8, value="")
                        ws.cell(row=current_row, column=9, value=daily_total_str)
                        
                        # Style the TOTAL row
                        for col in range(1, 10):
                            cell = ws.cell(row=current_row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        
                        current_row += 1
                        
            except Exception as e:
                self.logger.error(f"Error processing employee {emp_name}: {str(e)}")
                # Add error row
                ws.cell(row=current_row, column=1, value=emp_name)
                ws.cell(row=current_row, column=2, value=identity_type)
                ws.cell(row=current_row, column=3, value=identity_number)
                ws.cell(row=current_row, column=4, value="Error")
                ws.cell(row=current_row, column=5, value="Error")
                ws.cell(row=current_row, column=6, value="")
                ws.cell(row=current_row, column=7, value="Error")
                ws.cell(row=current_row, column=8, value="Error")
                ws.cell(row=current_row, column=9, value="00:00:00")
                current_row += 1
        
        return current_row
    
    def _generate_by_type_report(self, ws, employees, from_date, to_date, check_types_map, current_row):
        """Generate report grouped by type of fichaje"""
        # For now, use same logic as by_employee (can be enhanced later)
        return self._generate_by_employee_report(ws, employees, from_date, to_date, check_types_map, current_row)
    
    def _generate_by_group_report(self, ws, employees, from_date, to_date, check_types_map, current_row):
        """Generate report grouped by groups"""
        # For now, use same logic as by_employee (can be enhanced later)
        return self._generate_by_employee_report(ws, employees, from_date, to_date, check_types_map, current_row)

    def _create_empty_report(self) -> bytes:
        """Create an empty report when no data is found"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Vacío"
        
        ws.cell(row=1, column=1, value="No se encontraron empleados")
        ws.cell(row=2, column=1, value="Verifique los filtros aplicados")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_error_report(self, error_message: str) -> bytes:
        """Create an error report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Error"
        
        ws.cell(row=1, column=1, value="Error al generar el reporte")
        ws.cell(row=2, column=1, value=error_message)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_datetime(self, date_str: str) -> Optional[datetime]:
        """Parse datetime string from API"""
        if not date_str:
            return None
        
        try:
            return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        except ValueError:
            try:
                return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None

    def _get_employee_identification(self, employee: Dict) -> tuple:
        """Extract employee identification type and number"""
        # Try different possible field names for identification
        identification_number = (
            employee.get('nid') or
            employee.get('identificationNumber') or
            employee.get('code') or
            str(employee.get('pin')) if employee.get('pin') else None or
            employee.get('id', 'Sin número')
        )
        
        identification_type = (
            employee.get('identityNumberType') or
            employee.get('identificationType') or
            'DNI'
        )
        
        return identification_type, str(identification_number)
    
    def _process_break_redistribution(self, time_entries, break_entries):
        """Process break time redistribution based on workEntryType"""
        if not break_entries:
            return time_entries
        
        # Convert time entries to a mutable list and sort by start time
        processed_entries = []
        for entry in time_entries:
            work_in = entry.get('workEntryIn', {})
            if work_in and work_in.get('date'):
                try:
                    start_time = self._parse_datetime(work_in.get('date'))
                    if start_time:
                        processed_entries.append((start_time, entry))
                except Exception:
                    continue
        
        # Sort by start time
        processed_entries.sort(key=lambda x: x[0])
        
        # Group break entries by date for processing
        breaks_by_date = {}
        for break_entry in break_entries:
            if break_entry.get('date'):
                try:
                    break_date = self._parse_datetime(break_entry.get('date'))
                    if break_date:
                        date_key = break_date.strftime('%Y-%m-%d')
                        if date_key not in breaks_by_date:
                            breaks_by_date[date_key] = []
                        breaks_by_date[date_key].append(break_entry)
                except Exception:
                    continue
        
        # Process each date's entries
        final_entries = []
        current_date = None
        date_entries = []
        
        for start_time, entry in processed_entries:
            entry_date = start_time.strftime('%Y-%m-%d')
            
            # If we're on a new date, process the previous date's entries
            if current_date and current_date != entry_date:
                final_entries.extend(self._redistribute_break_time_for_date(
                    date_entries, breaks_by_date.get(current_date, [])
                ))
                date_entries = []
            
            current_date = entry_date
            date_entries.append(entry)
        
        # Process the last date's entries
        if date_entries:
            final_entries.extend(self._redistribute_break_time_for_date(
                date_entries, breaks_by_date.get(current_date, [])
            ))
        
        return final_entries
    
    def _redistribute_break_time_for_date(self, date_entries, date_breaks):
        """Redistribute break time for a specific date"""
        if not date_breaks:
            return date_entries
        
        # Filter out pause entries and collect their durations
        work_entries = []
        pause_duration_total = timedelta()
        
        for entry in date_entries:
            work_entry_type = entry.get('workEntryType', '').lower()
            
            # Check if this is a pause entry (contains "pausa", "descanso", "break", etc.)
            is_pause = any(pause_word in work_entry_type for pause_word in 
                          ['pausa', 'descanso', 'break', 'lunch', 'comida', 'almuerzo'])
            
            if is_pause:
                # Calculate pause duration and add to total
                work_in = entry.get('workEntryIn', {})
                work_out = entry.get('workEntryOut', {})
                if work_in and work_out:
                    try:
                        start_time = self._parse_datetime(work_in.get('date'))
                        end_time = self._parse_datetime(work_out.get('date'))
                        if start_time and end_time:
                            pause_duration_total += (end_time - start_time)
                    except Exception:
                        continue
                # Skip adding this entry to work_entries
            else:
                work_entries.append(entry)
        
        # If we have pause time to redistribute and work entries to distribute it to
        if pause_duration_total.total_seconds() > 0 and work_entries:
            # Distribute pause time evenly among work entries
            pause_per_entry = pause_duration_total.total_seconds() / len(work_entries)
            
            # Modify work entries to include redistributed time
            for entry in work_entries:
                work_out = entry.get('workEntryOut', {})
                if work_out and work_out.get('date'):
                    try:
                        original_end_time = self._parse_datetime(work_out.get('date'))
                        if original_end_time:
                            # Add redistributed pause time to the end time
                            new_end_time = original_end_time + timedelta(seconds=pause_per_entry)
                            # Update the workEntryOut date
                            entry['workEntryOut']['date'] = new_end_time.isoformat() + 'Z'
                    except Exception:
                        continue
        
        return work_entries