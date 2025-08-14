import openpyxl
import csv
import logging
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Union
from io import BytesIO, StringIO
from openpyxl.styles import Font, PatternFill, Alignment
from services.sesame_api import SesameAPI
from services.parallel_sesame_api import ParallelSesameAPI

class NoBreaksReportGenerator:
    def __init__(self):
        # Use parallel API for much faster processing
        self.sesame_api = ParallelSesameAPI()
        # Create a regular SesameAPI instance for collections mapping
        self.regular_api = SesameAPI()
        self.logger = logging.getLogger(__name__)
        # Cache for employee data and collections to avoid repeated processing
        self._employee_cache = {}
        self._collections_cache = None

    def generate_report(self, from_date: Optional[str] = None, to_date: Optional[str] = None, 
                       employee_id: Optional[str] = None, office_id: Optional[str] = None, 
                       department_id: Optional[str] = None, report_type: str = "by_employee", 
                       format: str = "xlsx", progress_callback = None, cancellation_token = None) -> Optional[bytes]:
        """Generate report with only work entries - no employee data processing"""
        
        try:
            self.logger.info(f"[REPORT] Starting report generation - from_date: {from_date}, to_date: {to_date}, report_type: {report_type}, format: {format}")
            
            # Ensure check types are cached
            from services.check_types_service import CheckTypesService
            check_types_service = CheckTypesService()
            self.logger.info("[REPORT] Ensuring check types are cached...")
            if not check_types_service.ensure_check_types_cached():
                self.logger.warning("Failed to cache check types, activity names may be incomplete")
            
            # OPTIMIZACIÓN: Cache collections mapping
            if self._collections_cache is None:
                self.logger.info("[REPORT] Fetching check type collections mapping...")
                self._collections_cache = self.regular_api.get_all_check_type_collections_mapping()
                self.logger.info(f"[REPORT] Collections mapping cached with {len(self._collections_cache)} check types")
            else:
                self.logger.info(f"[REPORT] Using cached collections mapping with {len(self._collections_cache)} check types")
            
            collections_mapping = self._collections_cache
            
            # OPTIMIZACIÓN: Usar procesamiento paralelo mejorado
            self.logger.info(f"[REPORT] Starting PARALLEL work entries retrieval...")
            start_time = time.time()
            
            all_work_entries = self.sesame_api.get_all_time_tracking_data_parallel(
                employee_id=employee_id,
                from_date=from_date,
                to_date=to_date,
                max_pages=100,
                max_workers=8  # Increased workers for better performance
            )
            
            fetch_time = time.time() - start_time
            self.logger.info(f"[REPORT] PARALLEL API fetch completed in {fetch_time:.1f}s - Total entries: {len(all_work_entries)}")
            
            # Check for cancellation after API fetch
            self._check_cancellation(cancellation_token, "after API fetch")
            
            # Simulate progress callback for compatibility
            if progress_callback and all_work_entries:
                progress_callback(1, 1, len(all_work_entries), len(all_work_entries))
            
            if not all_work_entries:
                return self._create_empty_report(format)

            self.logger.info(f"[REPORT] API fetch completed - Total entries retrieved: {len(all_work_entries)}")
            
            # Check for cancellation before cache warming
            self._check_cancellation(cancellation_token, "before cache warming")
            
            # OPTIMIZACIÓN: Pre-warm caches
            self.logger.info("[REPORT] Pre-warming caches for optimal performance...")
            self._warm_up_caches(all_work_entries)
            self.logger.info("[REPORT] Cache warm-up completed, starting report processing...")
            
            # Final check before intensive processing
            self._check_cancellation(cancellation_token, "before report processing")
            
            # Generate report based on format
            if format.lower() == "csv":
                return self._generate_csv_report(all_work_entries, collections_mapping, report_type, cancellation_token)
            else:
                return self._generate_xlsx_report(all_work_entries, collections_mapping, report_type, cancellation_token)
            
        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            return self._create_error_report(str(e), format)

    def _generate_xlsx_report(self, all_work_entries, collections_mapping, report_type, cancellation_token=None):
        """Generate XLSX report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set title based on report type
        if report_type == "by_group":
            ws.title = "Grupos y tipos de registro"
        else:
            ws.title = "Reporte Fichajes"
        
        # Headers based on report type
        if report_type == "by_group":
            headers = ["Grupo", "Actividad", "Fecha", "Empleado", "Tipo de identificación", "Nº de identificación", "Entrada", "Salida", "Tiempo registrado"]
        else:
            headers = ["Empleado", "Tipo ID", "Nº ID", "Fecha", "Actividad", "Grupo", "Entrada", "Salida", "Tiempo Registrado"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        current_row = 2
        
        # Check cancellation before processing
        self._check_cancellation(cancellation_token, "before XLSX processing")
        
        # Process entries based on report type
        if report_type == "by_activity":
            current_row = self._process_grouped_by_activity(ws, all_work_entries, collections_mapping, current_row, cancellation_token)
        elif report_type == "by_group":
            current_row = self._process_grouped_by_group(ws, all_work_entries, collections_mapping, current_row, cancellation_token)
        else:  # by_employee (default)
            current_row = self._process_grouped_entries(ws, all_work_entries, collections_mapping, current_row, cancellation_token)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _generate_csv_report(self, all_work_entries, collections_mapping, report_type, cancellation_token=None):
        """Generate CSV report"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Headers based on report type
        if report_type == "by_group":
            headers = ["Grupo", "Actividad", "Fecha", "Empleado", "Tipo de identificación", "Nº de identificación", "Entrada", "Salida", "Tiempo registrado"]
        else:
            headers = ["Empleado", "Tipo ID", "Nº ID", "Fecha", "Actividad", "Grupo", "Entrada", "Salida", "Tiempo Registrado"]
        writer.writerow(headers)
        
        # Process entries based on report type
        if report_type == "by_activity":
            self._process_grouped_by_activity_csv(writer, all_work_entries, collections_mapping)
        elif report_type == "by_group":
            self._process_grouped_by_group_csv(writer, all_work_entries, collections_mapping)
        else:  # by_employee (default)
            self._process_grouped_entries_csv(writer, all_work_entries, collections_mapping)
        
        # Convert to bytes
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8-sig')  # UTF-8 BOM for Excel compatibility

    def _create_empty_report(self, format: str = "xlsx") -> bytes:
        """Create an empty report when no data is found"""
        if format.lower() == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(["No se encontraron datos para los filtros especificados"])
            csv_content = output.getvalue()
            output.close()
            return csv_content.encode('utf-8-sig')
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Vacío"
            
            ws.cell(row=1, column=1, value="No se encontraron datos para los filtros especificados")
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

    def _create_error_report(self, error_message: str, format: str = "xlsx") -> bytes:
        """Create an error report"""
        if format.lower() == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow([f"Error al generar reporte: {error_message}"])
            csv_content = output.getvalue()
            output.close()
            return csv_content.encode('utf-8-sig')
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Error"
            
            ws.cell(row=1, column=1, value=f"Error al generar reporte: {error_message}")
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

    def _parse_datetime(self, date_str: str) -> Optional[datetime]:
        """Parse datetime string from API"""
        if not date_str:
            return None
        
        try:
            # Handle different datetime formats
            if date_str.endswith('Z'):
                date_str = date_str.replace('Z', '+00:00')
            
            return datetime.fromisoformat(date_str)
        except Exception:
            return None

    def _get_employee_identification(self, employee: Dict) -> tuple:
        """Extract employee identification type and number"""
        identification_type = employee.get('identityNumberType', 'DNI')
        identification_number = employee.get('nid', employee.get('id', 'No disponible'))
        
        return identification_type, identification_number

    def _redistribute_pause_time(self, entries: List[Dict]) -> List[Dict]:
        """Redistribute pause time by eliminating gaps and adjusting adjacent work entries"""
        if not entries:
            return entries
        
        processed_entries = []
        i = 0
        
        while i < len(entries):
            entry = entries[i]
            work_entry_type = entry.get('workEntryType', '')
            
            if work_entry_type == 'pause':
                # This is a pause entry - adjust adjacent work entries to eliminate gap
                pause_start = self._get_entry_start_time(entry)
                pause_end = self._get_entry_end_time(entry)
                
                if pause_start and pause_end:
                    # Find previous and next work entries
                    prev_entry = self._find_previous_work_entry(entries, i)
                    next_entry = self._find_next_work_entry(entries, i)
                    
                    if prev_entry:
                        # PRIORITY: Always try to extend previous entry to cover the pause
                        prev_start = self._get_entry_start_time(prev_entry)
                        prev_end = self._get_entry_end_time(prev_entry)
                        
                        if prev_start and prev_end:
                            # Extend previous entry to end of pause
                            self._extend_entry_to_time(prev_entry, pause_end)
                
                # Skip adding this pause entry to processed_entries
                i += 1
                continue
            else:
                # This is a work entry - add it to processed entries if not marked to skip
                if not entry.get('_skip', False):
                    processed_entries.append(entry)
                i += 1
        
        return processed_entries

    def _get_entry_duration_seconds(self, entry: Dict) -> int:
        """Get the duration of an entry in seconds"""
        try:
            work_entry_in = entry.get('workEntryIn', {})
            work_entry_out = entry.get('workEntryOut', {})
            
            if not work_entry_in.get('date') or not work_entry_out.get('date'):
                return 0
            
            in_time = datetime.fromisoformat(work_entry_in['date'].replace('Z', '+00:00'))
            out_time = datetime.fromisoformat(work_entry_out['date'].replace('Z', '+00:00'))
            
            duration = out_time - in_time
            return int(duration.total_seconds())
        except Exception as e:
            self.logger.error(f"Error calculating entry duration: {e}")
            return 0

    def _get_entry_start_time(self, entry: Dict) -> Optional[datetime]:
        """Get the start time of an entry"""
        try:
            work_entry_in = entry.get('workEntryIn', {})
            if work_entry_in.get('date'):
                return datetime.fromisoformat(work_entry_in['date'].replace('Z', '+00:00'))
        except Exception:
            pass
        return None

    def _get_entry_end_time(self, entry: Dict) -> Optional[datetime]:
        """Get the end time of an entry"""
        try:
            work_entry_out = entry.get('workEntryOut', {})
            if work_entry_out.get('date'):
                return datetime.fromisoformat(work_entry_out['date'].replace('Z', '+00:00'))
        except Exception:
            pass
        return None

    def _find_previous_work_entry(self, entries: List[Dict], pause_index: int) -> Optional[Dict]:
        """Find the previous work entry before the pause"""
        for i in range(pause_index - 1, -1, -1):
            if entries[i].get('workEntryType', '') != 'pause':
                return entries[i]
        return None

    def _find_next_work_entry(self, entries: List[Dict], pause_index: int) -> Optional[Dict]:
        """Find the next work entry after the pause"""
        for i in range(pause_index + 1, len(entries)):
            if entries[i].get('workEntryType', '') != 'pause':
                return entries[i]
        return None

    def _extend_entry_to_time(self, entry: Dict, end_time: datetime):
        """Extend a work entry to end at the specified time and update worked seconds"""
        try:
            work_entry_in = entry.get('workEntryIn', {})
            work_entry_out = entry.get('workEntryOut', {})
            
            if work_entry_in and work_entry_in.get('date') and work_entry_out:
                # Get the start time
                start_time = datetime.fromisoformat(work_entry_in['date'].replace('Z', '+00:00'))
                
                # Update end time
                work_entry_out['date'] = end_time.isoformat().replace('+00:00', 'Z')
                
                # Update worked seconds to reflect the new duration
                # Handle night shifts - if end_time appears before start_time, it's next day
                if end_time < start_time:
                    # Add 24 hours to end_time for calculation
                    adjusted_end_time = end_time + timedelta(days=1)
                    new_duration = adjusted_end_time - start_time
                else:
                    new_duration = end_time - start_time
                
                entry['workedSeconds'] = int(new_duration.total_seconds())
                
                # Format duration as HH:MM:SS
                hours = int(new_duration.total_seconds() // 3600)
                minutes = int((new_duration.total_seconds() % 3600) // 60)
                seconds = int(new_duration.total_seconds() % 60)
                duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                
                # Commented out to reduce log noise - uncomment for debugging
                # self.logger.debug(f"Extended entry: start {start_time.strftime('%H:%M:%S')}, new end {end_time.strftime('%H:%M:%S')}, new duration {duration_str} ({entry['workedSeconds']}s)")
        except Exception as e:
            self.logger.error(f"Error extending entry to time: {e}")

    def _move_entry_start_to_time(self, entry: Dict, start_time: datetime):
        """Move a work entry to start at the specified time and update worked seconds"""
        try:
            work_entry_in = entry.get('workEntryIn', {})
            work_entry_out = entry.get('workEntryOut', {})
            
            if work_entry_in:
                # Update start time
                work_entry_in['date'] = start_time.isoformat().replace('+00:00', 'Z')
                
                # Update worked seconds only if we have an end time
                if work_entry_out and work_entry_out.get('date'):
                    end_time = datetime.fromisoformat(work_entry_out['date'].replace('Z', '+00:00'))
                    new_duration = end_time - start_time
                    entry['workedSeconds'] = int(new_duration.total_seconds())

        except Exception as e:
            self.logger.error(f"Error moving entry start time: {e}")

    def _extend_entry_by_duration(self, entry: Dict, duration_seconds: int):
        """Extend a work entry by the specified duration"""
        try:
            work_entry_out = entry.get('workEntryOut', {})
            if work_entry_out and work_entry_out.get('date'):
                end_time = datetime.fromisoformat(work_entry_out['date'].replace('Z', '+00:00'))
                new_end_time = end_time + timedelta(seconds=duration_seconds)
                work_entry_out['date'] = new_end_time.isoformat().replace('+00:00', 'Z')
        except Exception as e:
            self.logger.error(f"Error extending entry by duration: {e}")

    def _get_entry_sort_key(self, entry: Dict):
        """Get sort key for chronological ordering by entry start time - handles night shifts"""
        try:
            work_entry_in = entry.get('workEntryIn', {})
            if work_entry_in and work_entry_in.get('date'):
                # Parse the datetime and return it for sorting
                parsed_time = datetime.fromisoformat(work_entry_in['date'].replace('Z', '+00:00'))
                
                # For night shifts: if time is between 00:00 and 06:00, add 24 hours for proper sorting
                # This ensures night shift entries (like 22:00, 23:00, 00:00, 01:00, 02:00) sort correctly
                if parsed_time.hour >= 0 and parsed_time.hour <= 6:
                    # This is likely early morning of next day in a night shift
                    # Add 24 hours to make it sort after the previous night's entries
                    sort_time = parsed_time + timedelta(hours=24)
                    return sort_time
                else:
                    return parsed_time
        except Exception as e:
            self.logger.error(f"Error parsing entry date for sorting: {e}")
        
        # Return a very old datetime as fallback for entries without valid dates
        return datetime.min.replace(tzinfo=datetime.now().astimezone().tzinfo)

    def _process_grouped_entries(self, ws, all_work_entries, collections_mapping, current_row):
        """Process entries grouped by employee and date, redistributing pause time"""
        # Group entries by employee and date
        grouped_entries = {}
        
        for entry in all_work_entries:
            # Get employee info
            employee_info = entry.get('employee', {})
            employee_name = f"{employee_info.get('firstName', '')} {employee_info.get('lastName', '')}".strip()
            employee_id = employee_info.get('id', 'unknown')
            
            if not employee_name:
                employee_name = "Empleado desconocido"
            
            # Extract date from workEntryIn.date
            entry_date = "No disponible"
            if entry.get('workEntryIn') and entry['workEntryIn'].get('date'):
                try:
                    entry_datetime = datetime.fromisoformat(
                        entry['workEntryIn']['date'].replace('Z', '+00:00'))
                    entry_date = entry_datetime.strftime('%d/%m/%Y')
                except Exception as e:
                    self.logger.error(f"Error parsing entry date: {e}")
                    entry_date = "Error en fecha"
            
            # Create group key by employee and date
            group_key = f"{employee_id}_{entry_date}"
            
            if group_key not in grouped_entries:
                grouped_entries[group_key] = {
                    'employee_name': employee_name,
                    'employee_id': employee_id,
                    'employee_info': employee_info,
                    'date': entry_date,
                    'all_entries': []
                }
            
            # Add ALL entries (work, pause, everything)
            grouped_entries[group_key]['all_entries'].append(entry)
        
        # Sort groups by employee name and date
        sorted_groups = sorted(grouped_entries.values(), 
                             key=lambda x: (x['employee_name'], x['date']))
        
        # Process each group
        for group in sorted_groups:
            all_entries = group['all_entries']
            
            # Sort all entries chronologically by entry start time
            all_entries.sort(key=self._get_entry_sort_key)
            
            # Process pause redistribution
            processed_entries = self._redistribute_pause_time(all_entries)
            
            # Sort processed entries again to ensure chronological order after pause redistribution
            processed_entries.sort(key=self._get_entry_sort_key)
            
            # Write processed entries to Excel (without pause entries)
            daily_totals = {}
            total_worked_seconds = 0
            
            # OPTIMIZACIÓN: Batch process entries for Excel writing
            batch_data = []
            for entry in processed_entries:
                row_data = self._extract_entry_data(entry, group['employee_info'], collections_mapping)
                batch_data.append([
                    row_data['employee_name'],
                    row_data['employee_id_type'],
                    row_data['employee_nid'],
                    row_data['entry_date'],
                    row_data['activity_name'],
                    row_data['group_name'],
                    row_data['start_time'],
                    row_data['end_time'],
                    row_data['final_duration']
                ])
                
                # Accumulate totals by activity type
                activity_name = row_data['activity_name']
                worked_seconds = row_data['worked_seconds']
                
                if activity_name not in daily_totals:
                    daily_totals[activity_name] = 0
                daily_totals[activity_name] += worked_seconds
                total_worked_seconds += worked_seconds
            
            # Write batch data to Excel - much faster than cell by cell
            for row_values in batch_data:
                for col_idx, value in enumerate(row_values, 1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
            
            # Add TOTAL row for this employee/date combination
            current_row = self._add_total_row(ws, group, daily_totals, total_worked_seconds, current_row)
            
            # Add blank row between different employee/date groups
            current_row += 1
        
        return current_row
    
    def _check_cancellation(self, cancellation_token, context=""):
        """Helper method to check cancellation and raise InterruptedError if cancelled"""
        if cancellation_token and cancellation_token.should_cancel():
            self.logger.info(f"[REPORT] Generation cancelled {context}")
            raise InterruptedError("Report generation was cancelled")
    
    def _warm_up_caches(self, all_work_entries):
        """Pre-warm all caches for optimal performance"""
        try:
            # Initialize check types service if not exists
            if not hasattr(self, '_check_types_service'):
                from services.check_types_service import CheckTypesService
                self._check_types_service = CheckTypesService()
            
            # Warm up activity cache
            self._check_types_service.warm_up_cache(all_work_entries)
            
            # Pre-cache unique employees
            unique_employees = {}
            for entry in all_work_entries:
                employee_info = entry.get('employee', {})
                employee_id = employee_info.get('id', 'unknown')
                if employee_id not in unique_employees:
                    unique_employees[employee_id] = employee_info
            
            # Pre-populate employee cache
            for emp_id, emp_info in unique_employees.items():
                if emp_id not in self._employee_cache:
                    employee_name = f"{emp_info.get('firstName', '')} {emp_info.get('lastName', '')}".strip()
                    if not employee_name:
                        employee_name = "Empleado desconocido"
                    
                    employee_nid = emp_info.get('nid', 'No disponible')
                    employee_id_type = emp_info.get('identityNumberType', 'DNI')
                    if employee_id_type:
                        employee_id_type = employee_id_type.lower()
                    
                    self._employee_cache[emp_id] = {
                        'name': employee_name,
                        'nid': employee_nid,
                        'id_type': employee_id_type
                    }
            
            self.logger.info(f"Caches warmed up - Employees: {len(self._employee_cache)}, Activities: {len(self._check_types_service._activity_cache)}")
            
        except Exception as e:
            self.logger.error(f"Error warming up caches: {str(e)}")
    
    def _extract_entry_data(self, entry, employee_info, collections_mapping=None):
        """Extract data from a work entry for Excel output - with caching"""
        # OPTIMIZACIÓN: Use employee cache
        employee_id = employee_info.get('id', 'unknown')
        
        if employee_id not in self._employee_cache:
            # Employee name
            employee_name = f"{employee_info.get('firstName', '')} {employee_info.get('lastName', '')}".strip()
            if not employee_name:
                employee_name = "Empleado desconocido"
            
            # Employee identification
            employee_nid = employee_info.get('nid', 'No disponible')
            employee_id_type = employee_info.get('identityNumberType', 'DNI')
            # Convert to lowercase to match the example format
            if employee_id_type:
                employee_id_type = employee_id_type.lower()
            
            # Cache employee data
            self._employee_cache[employee_id] = {
                'name': employee_name,
                'nid': employee_nid,
                'id_type': employee_id_type
            }
        
        # Get cached employee data
        cached_employee = self._employee_cache[employee_id]
        employee_name = cached_employee['name']
        employee_nid = cached_employee['nid']
        employee_id_type = cached_employee['id_type']
        
        # OPTIMIZACIÓN: Use cached check types service
        work_entry_type = entry.get('workEntryType', '')
        work_break_id = entry.get('workBreakId')
        
        # Use the cached instance
        activity_name = self._check_types_service.get_activity_name(work_entry_type, work_break_id)
        
        # Get group name from collections mapping using workCheckTypeId
        work_check_type_id = entry.get('workCheckTypeId')
        group_name = "Sin Grupo"
        if collections_mapping and work_check_type_id:
            group_name = collections_mapping.get(work_check_type_id, "Sin Grupo")
            self.logger.debug(f"Work entry with check_type_id {work_check_type_id} mapped to group: {group_name}")
        
        # Extract date from workEntryIn.date
        entry_date = "No disponible"
        if entry.get('workEntryIn') and entry['workEntryIn'].get('date'):
            try:
                entry_datetime = datetime.fromisoformat(
                    entry['workEntryIn']['date'].replace('Z', '+00:00'))
                entry_date = entry_datetime.strftime('%d/%m/%Y')
            except Exception as e:
                self.logger.error(f"Error parsing entry date: {e}")
                entry_date = "Error en fecha"
        
        # Extract times from workEntryIn and workEntryOut
        start_time = "No disponible"
        end_time = "No disponible"
        
        if entry.get('workEntryIn') and entry['workEntryIn'].get('date'):
            try:
                start_datetime = datetime.fromisoformat(
                    entry['workEntryIn']['date'].replace('Z', '+00:00'))
                start_time = start_datetime.strftime('%H:%M:%S')
            except Exception as e:
                self.logger.error(f"Error parsing start time: {e}")
                start_time = "Error en hora"
        
        if entry.get('workEntryOut') and entry['workEntryOut'].get('date'):
            try:
                end_datetime = datetime.fromisoformat(
                    entry['workEntryOut']['date'].replace('Z', '+00:00'))
                end_time = end_datetime.strftime('%H:%M:%S')
            except Exception as e:
                self.logger.error(f"Error parsing end time: {e}")
                end_time = "Error en hora"
        
        # Calculate duration
        worked_seconds = entry.get('workedSeconds', 0)
        final_duration = self._format_duration(timedelta(seconds=worked_seconds))
        
        return {
            'employee_name': employee_name,
            'employee_id_type': employee_id_type,
            'employee_nid': employee_nid,
            'entry_date': entry_date,
            'activity_name': activity_name,
            'group_name': group_name,
            'start_time': start_time,
            'end_time': end_time,
            'final_duration': final_duration,
            'worked_seconds': worked_seconds
        }
    
    def _add_total_row(self, ws, group, daily_totals, total_worked_seconds, current_row):
        """Add TOTAL row for employee/date combination"""
        employee_info = group['employee_info']
        employee_name = group['employee_name']
        employee_nid = employee_info.get('nid', 'No disponible')
        employee_id_type = employee_info.get('identityNumberType', 'DNI')
        entry_date = group['date']
        
        # Create summary of activity types
        activity_summary = []
        for activity_name, seconds in daily_totals.items():
            duration_str = self._format_duration(timedelta(seconds=seconds))
            activity_summary.append(f"{activity_name}: {duration_str}")
        
        # Join activity summary (limit to avoid Excel cell size issues)
        activity_summary_str = "; ".join(activity_summary[:5])  # Limit to 5 activities
        if len(daily_totals) > 5:
            activity_summary_str += f"; ... y {len(daily_totals) - 5} más"
        
        # Total duration
        total_duration = self._format_duration(timedelta(seconds=total_worked_seconds))
        
        # Apply bold formatting for TOTAL row
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Write TOTAL row
        ws.cell(row=current_row, column=1, value=employee_name).font = total_font
        ws.cell(row=current_row, column=2, value=employee_id_type).font = total_font
        ws.cell(row=current_row, column=3, value=employee_nid).font = total_font
        ws.cell(row=current_row, column=4, value=entry_date).font = total_font
        ws.cell(row=current_row, column=5, value="TOTAL").font = total_font
        ws.cell(row=current_row, column=6, value="").font = total_font
        ws.cell(row=current_row, column=7, value="").font = total_font
        ws.cell(row=current_row, column=8, value="").font = total_font
        ws.cell(row=current_row, column=9, value=total_duration).font = total_font
        
        # Apply background color to TOTAL row
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).fill = total_fill
        
        return current_row + 1

    def _format_duration(self, duration):
        """Format duration as HH:MM:SS - same as preview"""
        if isinstance(duration, timedelta):
            total_seconds = int(duration.total_seconds())
        else:
            total_seconds = int(duration)

        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60

        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def _process_grouped_entries_csv(self, writer, all_work_entries, collections_mapping):
        """Process entries grouped by employee and date for CSV output"""
        # Reuse the Excel logic but write to CSV
        import csv
        from io import StringIO
        
        # Create a temporary workbook-like structure for reusing existing logic
        class CSVRow:
            def __init__(self, writer):
                self.writer = writer
                self.row_data = []
            
            def cell(self, row, column, value=None):
                # Adjust row_data size if needed
                while len(self.row_data) < column:
                    self.row_data.append("")
                if value is not None:
                    self.row_data[column-1] = value
                return self
        
        # Create a mock worksheet that writes to CSV
        class CSVWorksheet:
            def __init__(self, writer):
                self.writer = writer
                self.current_row = 0
            
            def cell(self, row, column, value=None):
                if row > self.current_row:
                    # Write accumulated row data and start new row
                    if hasattr(self, 'row_data') and self.row_data:
                        self.writer.writerow(self.row_data)
                    self.row_data = [""] * 10  # Initialize with empty cells
                    self.current_row = row
                
                if value is not None:
                    while len(self.row_data) < column:
                        self.row_data.append("")
                    self.row_data[column-1] = str(value)
                
                return self
        
        # Create CSV worksheet wrapper
        csv_ws = CSVWorksheet(writer)
        
        # Reuse existing Excel processing logic
        current_row = self._process_grouped_entries(csv_ws, all_work_entries, collections_mapping, 1)

    def _process_grouped_by_activity_csv(self, writer, all_work_entries, collections_mapping):
        """Process entries grouped by activity type for CSV output"""
        # Group entries by activity type first
        activity_groups = {}
        
        for entry in all_work_entries:
            # Get activity name
            work_entry_type = entry.get('workEntryType', '')
            work_break_id = entry.get('workBreakId')
            
            from services.check_types_service import CheckTypesService
            check_types_service = CheckTypesService()
            activity_name = check_types_service.get_activity_name(work_entry_type, work_break_id)
            
            if activity_name not in activity_groups:
                activity_groups[activity_name] = []
            activity_groups[activity_name].append(entry)
        
        # Process each activity group separately
        for activity_name in sorted(activity_groups.keys()):
            entries = activity_groups[activity_name]
            
            # Write activity header row
            writer.writerow([f"=== ACTIVIDAD: {activity_name} ===", "", "", "", "", "", "", "", ""])
            
            # Process entries for this activity using employee grouping logic
            self._process_grouped_entries_csv(writer, entries, collections_mapping)
            
            # Add separator between activity groups
            writer.writerow([])

    def _process_grouped_by_group_csv(self, writer, all_work_entries, collections_mapping):
        """Process entries grouped by groups for CSV output"""
        # First, create a list of all entries with their group names and dates
        entries_with_groups = []
        
        for entry in all_work_entries:
            # Get group name from collections mapping
            work_check_type_id = entry.get('workCheckTypeId')
            group_name = "Sin Grupo"
            if collections_mapping and work_check_type_id:
                group_name = collections_mapping.get(work_check_type_id, "Sin Grupo")
            
            # Extract date from entry
            entry_date = "No disponible"
            if entry.get('workEntryIn') and entry['workEntryIn'].get('date'):
                try:
                    entry_datetime = datetime.fromisoformat(
                        entry['workEntryIn']['date'].replace('Z', '+00:00'))
                    entry_date = entry_datetime.strftime('%d/%m/%Y')
                except Exception as e:
                    self.logger.error(f"Error parsing entry date: {e}")
            
            # Store entry with its group name and date
            entries_with_groups.append({
                'group_name': group_name,
                'entry_date': entry_date,
                'entry': entry
            })
        
        # Sort entries first by group name, then by date, then by time
        def get_combined_sort_key(item):
            group_name = item['group_name']
            entry_date = item['entry_date']
            entry_datetime = self._get_entry_sort_key(item['entry'])
            return (group_name, entry_date, entry_datetime)
        
        entries_with_groups.sort(key=get_combined_sort_key)
        
        # Group entries for pause redistribution by group AND date
        grouped_entries = {}
        for item in entries_with_groups:
            group_key = f"{item['group_name']}_{item['entry_date']}"
            if group_key not in grouped_entries:
                grouped_entries[group_key] = {
                    'group_name': item['group_name'],
                    'entry_date': item['entry_date'],
                    'entries': []
                }
            grouped_entries[group_key]['entries'].append(item['entry'])
        
        # Process each group/date combination for pause redistribution
        all_processed_entries = []
        for group_key, group_data in grouped_entries.items():
            # Process pause redistribution for this group/date
            processed_entries = self._redistribute_pause_time(group_data['entries'])
            
            # Add group name and date to each processed entry
            for entry in processed_entries:
                all_processed_entries.append({
                    'group_name': group_data['group_name'],
                    'entry_date': group_data['entry_date'],
                    'entry': entry
                })
        
        # Sort all processed entries again by group, date and time
        all_processed_entries.sort(key=get_combined_sort_key)
        
        # Write all entries to CSV with subtotals for each group/date
        current_group = None
        current_date = None
        group_date_total_seconds = 0
        
        for i, item in enumerate(all_processed_entries):
            group_name = item['group_name']
            entry_date = item['entry_date']
            entry = item['entry']
            
            # Check if we need to add a subtotal (group or date changed)
            if current_group is not None and (current_group != group_name or current_date != entry_date):
                # Add total row for the previous group/date combination
                if group_date_total_seconds > 0:
                    total_duration = self._format_duration(timedelta(seconds=group_date_total_seconds))
                    
                    # Write TOTAL row
                    writer.writerow([
                        current_group,  # Grupo
                        "TOTAL",  # Actividad
                        current_date,  # Fecha
                        "",  # Empleado
                        "",  # Tipo de identificación
                        "",  # Nº de identificación
                        "",  # Entrada
                        "",  # Salida
                        total_duration  # Tiempo registrado
                    ])
                    
                    # Add blank row after total
                    writer.writerow([])
                
                # Reset totals for new group/date
                group_date_total_seconds = 0
            
            # Update current group and date
            current_group = group_name
            current_date = entry_date
            
            # Get employee info for this entry
            employee_info = entry.get('employee', {})
            row_data = self._extract_entry_data(entry, employee_info, collections_mapping)
            
            # Override group name with our group (this ensures the grouped value is used)
            row_data['group_name'] = group_name
            
            # Write CSV row with columns in correct order
            writer.writerow([
                row_data['group_name'],  # Grupo
                row_data['activity_name'],  # Actividad
                row_data['entry_date'],  # Fecha
                row_data['employee_name'],  # Empleado
                row_data['employee_id_type'],  # Tipo de identificación
                row_data['employee_nid'],  # Nº de identificación
                row_data['start_time'],  # Entrada
                row_data['end_time'],  # Salida
                row_data['final_duration']  # Tiempo registrado
            ])
            
            # Accumulate totals
            worked_seconds = row_data['worked_seconds']
            group_date_total_seconds += worked_seconds
        
        # Add final total for the last group/date if exists
        if current_group is not None and group_date_total_seconds > 0:
            total_duration = self._format_duration(timedelta(seconds=group_date_total_seconds))
            
            # Write TOTAL row
            writer.writerow([
                current_group,  # Grupo
                "TOTAL",  # Actividad
                current_date,  # Fecha
                "",  # Empleado
                "",  # Tipo de identificación
                "",  # Nº de identificación
                "",  # Entrada
                "",  # Salida
                total_duration  # Tiempo registrado
            ])
    
    def get_data_metrics(self, from_date: Optional[str] = None, to_date: Optional[str] = None, 
                         employee_id: Optional[str] = None, office_id: Optional[str] = None, 
                         department_id: Optional[str] = None) -> dict:
        """Get data collection metrics without generating full report"""
        try:
            # Get all work entries
            all_work_entries = self.sesame_api.get_all_time_tracking_data(
                employee_id=employee_id,
                from_date=from_date,
                to_date=to_date
            )
            
            # Get check types mapping
            check_types_response = self.sesame_api.get_check_types()
            check_types_map = {}
            if check_types_response and 'data' in check_types_response:
                for check_type in check_types_response['data']:
                    check_types_map[check_type['id']] = check_type['name']
            
            # Process first 10 entries for preview
            preview_entries = []
            for entry in all_work_entries[:10]:  # Only first 10 for preview
                employee_info = entry.get('employee', {})
                employee_name = f"{employee_info.get('firstName', '')} {employee_info.get('lastName', '')}".strip()
                
                if not employee_name:
                    employee_name = "Empleado desconocido"
                
                # Extract date
                entry_date = "No disponible"
                if entry.get('workEntryIn') and entry['workEntryIn'].get('date'):
                    try:
                        entry_datetime = datetime.fromisoformat(
                            entry['workEntryIn']['date'].replace('Z', '+00:00'))
                        entry_date = entry_datetime.strftime('%Y-%m-%d')
                    except Exception:
                        entry_date = "Error en fecha"
                
                row_data = self._extract_entry_data(entry, employee_info)
                preview_entries.append(row_data)
            
            return {
                'total_entries': len(all_work_entries),
                'preview_entries': preview_entries,
                'success': True
            }
            
        except Exception as e:
            self.logger.error(f"Error getting data metrics: {str(e)}")
            return {
                'total_entries': 0,
                'preview_entries': [],
                'success': False,
                'error': str(e)
            }

    def _process_grouped_by_activity(self, ws, all_work_entries, collections_mapping, current_row):
        """Process entries grouped by activity type"""
        # Group entries by activity type and date
        grouped_entries = {}
        
        for entry in all_work_entries:
            # Get activity name based on workEntryType and workBreakId
            work_entry_type = entry.get('workEntryType', '')
            work_break_id = entry.get('workBreakId')
            
            # Import here to avoid circular import
            from services.check_types_service import CheckTypesService
            check_types_service = CheckTypesService()
            activity_name = check_types_service.get_activity_name(work_entry_type, work_break_id)
            
            # Extract date from workEntryIn.date
            entry_date = "No disponible"
            if entry.get('workEntryIn') and entry['workEntryIn'].get('date'):
                try:
                    entry_datetime = datetime.fromisoformat(
                        entry['workEntryIn']['date'].replace('Z', '+00:00'))
                    entry_date = entry_datetime.strftime('%d/%m/%Y')
                except Exception as e:
                    self.logger.error(f"Error parsing entry date: {e}")
                    entry_date = "Error en fecha"
            
            # Create group key by activity and date
            group_key = f"{activity_name}_{entry_date}"
            
            if group_key not in grouped_entries:
                grouped_entries[group_key] = {
                    'activity_name': activity_name,
                    'date': entry_date,
                    'all_entries': []
                }
            
            grouped_entries[group_key]['all_entries'].append(entry)
        
        # Sort groups by activity name and date
        sorted_groups = sorted(grouped_entries.values(), 
                             key=lambda x: (x['activity_name'], x['date']))
        
        # Process each group
        for group in sorted_groups:
            all_entries = group['all_entries']
            
            # Sort all entries chronologically by entry start time
            all_entries.sort(key=self._get_entry_sort_key)
            
            # Process pause redistribution
            processed_entries = self._redistribute_pause_time(all_entries)
            
            # Sort processed entries again to ensure chronological order after pause redistribution
            processed_entries.sort(key=self._get_entry_sort_key)
            
            # Write processed entries to Excel (without pause entries)
            activity_totals = {}
            total_worked_seconds = 0
            
            for entry in processed_entries:
                # Get employee info for this entry
                employee_info = entry.get('employee', {})
                row_data = self._extract_entry_data(entry, employee_info, collections_mapping)
                
                # Write to Excel
                ws.cell(row=current_row, column=1, value=row_data['employee_name'])
                ws.cell(row=current_row, column=2, value=row_data['employee_id_type'])
                ws.cell(row=current_row, column=3, value=row_data['employee_nid'])
                ws.cell(row=current_row, column=4, value=row_data['entry_date'])
                ws.cell(row=current_row, column=5, value=row_data['activity_name'])
                ws.cell(row=current_row, column=6, value=row_data['group_name'])
                ws.cell(row=current_row, column=7, value=row_data['start_time'])
                ws.cell(row=current_row, column=8, value=row_data['end_time'])
                ws.cell(row=current_row, column=9, value=row_data['final_duration'])
                
                # Accumulate totals
                worked_seconds = row_data['worked_seconds']
                total_worked_seconds += worked_seconds
                
                current_row += 1
            
            # Add TOTAL row for this activity/date combination
            current_row = self._add_activity_total_row(ws, group, total_worked_seconds, current_row)
            
            # Add blank row between different activity/date groups
            current_row += 1
        
        return current_row

    def _process_grouped_by_group(self, ws, all_work_entries, collections_mapping, current_row):
        """Process entries grouped by work groups using check type collections"""
        # First, create a list of all entries with their group names and dates
        entries_with_groups = []
        
        for entry in all_work_entries:
            # Get group name from collections mapping based on workCheckTypeId
            work_check_type_id = entry.get('workCheckTypeId')
            group_name = "Sin Grupo"
            if collections_mapping and work_check_type_id:
                group_name = collections_mapping.get(work_check_type_id, "Sin Grupo")
            
            # Extract date from entry
            entry_date = "No disponible"
            if entry.get('workEntryIn') and entry['workEntryIn'].get('date'):
                try:
                    entry_datetime = datetime.fromisoformat(
                        entry['workEntryIn']['date'].replace('Z', '+00:00'))
                    entry_date = entry_datetime.strftime('%d/%m/%Y')
                except Exception as e:
                    self.logger.error(f"Error parsing entry date: {e}")
            
            # Store entry with its group name and date
            entries_with_groups.append({
                'group_name': group_name,
                'entry_date': entry_date,
                'entry': entry
            })
        
        # Sort entries first by group name, then by date, then by time
        def get_combined_sort_key(item):
            group_name = item['group_name']
            entry_date = item['entry_date']
            entry_datetime = self._get_entry_sort_key(item['entry'])
            return (group_name, entry_date, entry_datetime)
        
        entries_with_groups.sort(key=get_combined_sort_key)
        
        # Group entries for pause redistribution by group AND date
        grouped_entries = {}
        for item in entries_with_groups:
            group_key = f"{item['group_name']}_{item['entry_date']}"
            if group_key not in grouped_entries:
                grouped_entries[group_key] = {
                    'group_name': item['group_name'],
                    'entry_date': item['entry_date'],
                    'entries': []
                }
            grouped_entries[group_key]['entries'].append(item['entry'])
        
        # Process each group/date combination for pause redistribution
        all_processed_entries = []
        for group_key, group_data in grouped_entries.items():
            # Process pause redistribution for this group/date
            processed_entries = self._redistribute_pause_time(group_data['entries'])
            
            # Add group name and date to each processed entry
            for entry in processed_entries:
                all_processed_entries.append({
                    'group_name': group_data['group_name'],
                    'entry_date': group_data['entry_date'],
                    'entry': entry
                })
        
        # Sort all processed entries again by group, date and time
        all_processed_entries.sort(key=get_combined_sort_key)
        
        # Write all entries to Excel with subtotals for each group/date
        current_group = None
        current_date = None
        group_date_total_seconds = 0
        
        for i, item in enumerate(all_processed_entries):
            group_name = item['group_name']
            entry_date = item['entry_date']
            entry = item['entry']
            
            # Check if we need to add a subtotal (group or date changed)
            if current_group is not None and (current_group != group_name or current_date != entry_date):
                # Add total row for the previous group/date combination
                if group_date_total_seconds > 0:
                    total_duration = self._format_duration(timedelta(seconds=group_date_total_seconds))
                    
                    # Apply bold formatting for TOTAL row
                    total_font = Font(bold=True)
                    total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    
                    # Write TOTAL row with same format as data rows
                    ws.cell(row=current_row, column=1, value=current_group).font = total_font
                    ws.cell(row=current_row, column=2, value="TOTAL").font = total_font
                    ws.cell(row=current_row, column=3, value=current_date).font = total_font
                    ws.cell(row=current_row, column=4, value="").font = total_font
                    ws.cell(row=current_row, column=5, value="").font = total_font
                    ws.cell(row=current_row, column=6, value="").font = total_font
                    ws.cell(row=current_row, column=7, value="").font = total_font
                    ws.cell(row=current_row, column=8, value="").font = total_font
                    ws.cell(row=current_row, column=9, value=total_duration).font = total_font
                    
                    # Apply background color to TOTAL row
                    for col in range(1, 10):
                        ws.cell(row=current_row, column=col).fill = total_fill
                    
                    current_row += 1
                    # Add a blank row after total
                    current_row += 1
                
                # Reset totals for new group/date
                group_date_total_seconds = 0
            
            # Update current group and date
            current_group = group_name
            current_date = entry_date
            
            # Get employee info for this entry
            employee_info = entry.get('employee', {})
            row_data = self._extract_entry_data(entry, employee_info, collections_mapping)
            
            # Override group name with our group (this ensures the grouped value is used)
            row_data['group_name'] = group_name
            
            # Write to Excel with columns: Grupo, Actividad, Fecha, Empleado, Tipo Doc, NID, Entrada, Salida, Duración
            ws.cell(row=current_row, column=1, value=row_data['group_name'])
            ws.cell(row=current_row, column=2, value=row_data['activity_name'])
            ws.cell(row=current_row, column=3, value=row_data['entry_date'])
            ws.cell(row=current_row, column=4, value=row_data['employee_name'])
            ws.cell(row=current_row, column=5, value=row_data['employee_id_type'])
            ws.cell(row=current_row, column=6, value=row_data['employee_nid'])
            ws.cell(row=current_row, column=7, value=row_data['start_time'])
            ws.cell(row=current_row, column=8, value=row_data['end_time'])
            ws.cell(row=current_row, column=9, value=row_data['final_duration'])
            
            # Accumulate totals
            worked_seconds = row_data['worked_seconds']
            group_date_total_seconds += worked_seconds
            
            current_row += 1
        
        # Add final total for the last group/date if exists
        if current_group is not None and group_date_total_seconds > 0:
            total_duration = self._format_duration(timedelta(seconds=group_date_total_seconds))
            
            # Apply bold formatting for TOTAL row
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # Write TOTAL row
            ws.cell(row=current_row, column=1, value=current_group).font = total_font
            ws.cell(row=current_row, column=2, value="TOTAL").font = total_font
            ws.cell(row=current_row, column=3, value=current_date).font = total_font
            ws.cell(row=current_row, column=4, value="").font = total_font
            ws.cell(row=current_row, column=5, value="").font = total_font
            ws.cell(row=current_row, column=6, value="").font = total_font
            ws.cell(row=current_row, column=7, value="").font = total_font
            ws.cell(row=current_row, column=8, value="").font = total_font
            ws.cell(row=current_row, column=9, value=total_duration).font = total_font
            
            # Apply background color to TOTAL row
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).fill = total_fill
            
            current_row += 1
        
        return current_row

    def _add_activity_total_row(self, ws, group, total_worked_seconds, current_row):
        """Add TOTAL row for activity/date combination"""
        activity_name = group['activity_name']
        entry_date = group['date']
        
        # Total duration
        total_duration = self._format_duration(timedelta(seconds=total_worked_seconds))
        
        # Apply bold formatting for TOTAL row
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Write TOTAL row
        ws.cell(row=current_row, column=1, value="TOTAL").font = total_font
        ws.cell(row=current_row, column=2, value="").font = total_font
        ws.cell(row=current_row, column=3, value="").font = total_font
        ws.cell(row=current_row, column=4, value=entry_date).font = total_font
        ws.cell(row=current_row, column=5, value=activity_name).font = total_font
        ws.cell(row=current_row, column=6, value="").font = total_font
        ws.cell(row=current_row, column=7, value="").font = total_font
        ws.cell(row=current_row, column=8, value="").font = total_font
        ws.cell(row=current_row, column=9, value=total_duration).font = total_font
        
        # Apply background color to TOTAL row
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).fill = total_fill
        
        return current_row + 1

    def _add_group_total_row(self, ws, group, total_worked_seconds, current_row):
        """Add TOTAL row for group"""
        group_name = group['group_name']
        
        # Total duration
        total_duration = self._format_duration(timedelta(seconds=total_worked_seconds))
        
        # Apply bold formatting for TOTAL row
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Write TOTAL row (columns: Grupo, Actividad, Fecha, Empleado, Tipo Doc, NID, Entrada, Salida, Duración)
        ws.cell(row=current_row, column=1, value=group_name).font = total_font
        ws.cell(row=current_row, column=2, value="TOTAL").font = total_font
        ws.cell(row=current_row, column=3, value="").font = total_font
        ws.cell(row=current_row, column=4, value="").font = total_font
        ws.cell(row=current_row, column=5, value="").font = total_font
        ws.cell(row=current_row, column=6, value="").font = total_font
        ws.cell(row=current_row, column=7, value="").font = total_font
        ws.cell(row=current_row, column=8, value="").font = total_font
        ws.cell(row=current_row, column=9, value=total_duration).font = total_font
        
        # Apply background color to TOTAL row
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).fill = total_fill
        
        return current_row + 1