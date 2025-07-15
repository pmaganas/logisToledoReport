import openpyxl
import logging
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from services.sesame_api import SesameAPI

class DebugReportGenerator:
    def __init__(self):
        self.sesame_api = SesameAPI()
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)
        self.api_calls = 0
        self.start_time = time.time()

    def log_api_call(self, endpoint: str, employee_name: str = None):
        """Log each API call with timing"""
        self.api_calls += 1
        elapsed = time.time() - self.start_time
        context = f" for {employee_name}" if employee_name else ""
        self.logger.info(f"API CALL #{self.api_calls}: {endpoint}{context} [Total time: {elapsed:.1f}s]")

    def generate_debug_report(self, from_date: str = None, to_date: str = None, 
                             employee_id: str = None, office_id: str = None, 
                             department_id: str = None, report_type: str = "by_employee") -> Optional[bytes]:
        """Generate report with detailed debugging info"""
        
        try:
            self.logger.info("=== INICIANDO GENERACIÓN DE REPORTE DEBUG ===")
            self.logger.info(f"Parámetros: from_date={from_date}, to_date={to_date}, employee_id={employee_id}")
            
            # Step 1: Get employees - limit to 3 for testing
            self.logger.info("PASO 1: Obteniendo empleados...")
            if employee_id:
                self.log_api_call("get_employee_details", employee_id)
                employee_response = self.sesame_api.get_employee_details(employee_id)
                employees = [employee_response.get('data')] if employee_response else []
            else:
                self.log_api_call("get_employees")
                employees_response = self.sesame_api.get_employees(page=1, per_page=100)
                all_employees = employees_response.get('data', []) if employees_response else []
                
                # Apply filters
                employees = []
                for employee in all_employees:
                    include_employee = True
                    
                    if office_id:
                        employee_office_id = employee.get('office', {}).get('id') if employee.get('office') else None
                        if employee_office_id != office_id:
                            include_employee = False
                    
                    if department_id and include_employee:
                        employee_department_id = employee.get('department', {}).get('id') if employee.get('department') else None
                        if employee_department_id != department_id:
                            include_employee = False
                    
                    if include_employee:
                        employees.append(employee)
                
                # LIMIT TO 3 EMPLOYEES FOR TESTING
                employees = employees[:3]
            
            if not employees:
                self.logger.warning("No employees found")
                return self._create_empty_report()
            
            self.logger.info(f"✓ Encontrados {len(employees)} empleados para procesar")
            
            # Step 2: Get check types once
            self.logger.info("PASO 2: Obteniendo tipos de actividad...")
            self.log_api_call("get_check_types")
            check_types_response = self.sesame_api.get_check_types(page=1, per_page=100)
            check_types_map = {}
            if check_types_response and check_types_response.get('data'):
                for check_type in check_types_response['data']:
                    check_types_map[check_type.get('id')] = check_type.get('name', 'Actividad no especificada')
                self.logger.info(f"✓ Cargados {len(check_types_map)} tipos de actividad")
            
            # Step 3: Process each employee
            self.logger.info("PASO 3: Procesando empleados...")
            all_employee_data = {}
            
            for i, employee in enumerate(employees, 1):
                emp_id = employee.get('id')
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                
                self.logger.info(f"--- Procesando empleado {i}/{len(employees)}: {emp_name} ---")
                
                # Get time entries
                self.log_api_call("get_work_entries", emp_name)
                time_response = self.sesame_api.get_work_entries(
                    employee_id=emp_id,
                    from_date=from_date,
                    to_date=to_date,
                    page=1,
                    limit=100
                )
                
                time_entries = []
                if time_response and time_response.get('data'):
                    time_entries = time_response['data']
                    self.logger.info(f"  ✓ {len(time_entries)} time entries")
                else:
                    self.logger.info(f"  ✗ No time entries")
                
                # Get break entries
                self.log_api_call("get_breaks", emp_name)
                break_response = self.sesame_api.get_breaks(
                    employee_id=emp_id,
                    from_date=from_date,
                    to_date=to_date,
                    page=1,
                    limit=100
                )
                
                break_entries = []
                if break_response and break_response.get('data'):
                    break_entries = break_response['data']
                    self.logger.info(f"  ✓ {len(break_entries)} break entries")
                else:
                    self.logger.info(f"  ✗ No break entries")
                
                all_employee_data[emp_id] = {
                    'employee': employee,
                    'time_entries': time_entries,
                    'break_entries': break_entries
                }
            
            # Step 4: Create Excel report
            self.logger.info("PASO 4: Generando archivo Excel...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Debug"
            
            # Headers
            headers = ["Empleado", "Tipo ID", "Número ID", "Fecha", "Actividad", "Grupo", "Entrada", "Salida", "Tiempo Registrado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            current_row = 2
            
            # Process employee data
            for emp_id, data in all_employee_data.items():
                employee = data['employee']
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                identity_type, identity_number = self._get_employee_identification(employee)
                
                time_entries = data['time_entries']
                
                if time_entries:
                    # Group by date
                    entries_by_date = {}
                    for entry in time_entries:
                        work_in = entry.get('workEntryIn', {}) if entry.get('workEntryIn') else {}
                        if work_in and work_in.get('date'):
                            try:
                                parsed_date = self._parse_datetime(work_in.get('date'))
                                if parsed_date:
                                    date_key = parsed_date.strftime('%Y-%m-%d')
                                    if date_key not in entries_by_date:
                                        entries_by_date[date_key] = []
                                    entries_by_date[date_key].append(entry)
                            except Exception:
                                continue
                    
                    # Process each date
                    for date_key in sorted(entries_by_date.keys()):
                        date_entries = entries_by_date[date_key]
                        daily_total_seconds = 0
                        
                        for entry in date_entries:
                            work_in = entry.get('workEntryIn', {})
                            work_out = entry.get('workEntryOut', {}) if entry.get('workEntryOut') else {}
                            
                            start_time = self._parse_datetime(work_in.get('date')) if work_in else None
                            end_time = self._parse_datetime(work_out.get('date')) if work_out else None
                            
                            # Calculate duration
                            duration_str = "00:00:00"
                            if start_time and end_time:
                                duration = end_time - start_time
                                entry_seconds = duration.total_seconds()
                                daily_total_seconds += entry_seconds
                                hours = int(entry_seconds // 3600)
                                minutes = int((entry_seconds % 3600) // 60)
                                seconds = int(entry_seconds % 60)
                                duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                            
                            # Get activity name
                            activity_name = entry.get('workEntryType', 'Trabajo')
                            
                            # Add row
                            ws.cell(row=current_row, column=1, value=emp_name)
                            ws.cell(row=current_row, column=2, value=identity_type)
                            ws.cell(row=current_row, column=3, value=identity_number)
                            ws.cell(row=current_row, column=4, value=date_key)
                            ws.cell(row=current_row, column=5, value=activity_name)
                            ws.cell(row=current_row, column=6, value="")
                            ws.cell(row=current_row, column=7, value=start_time.strftime("%H:%M:%S") if start_time else "N/A")
                            ws.cell(row=current_row, column=8, value=end_time.strftime("%H:%M:%S") if end_time else "N/A")
                            ws.cell(row=current_row, column=9, value=duration_str)
                            current_row += 1
                else:
                    # No time entries
                    ws.cell(row=current_row, column=1, value=emp_name)
                    ws.cell(row=current_row, column=2, value=identity_type)
                    ws.cell(row=current_row, column=3, value=identity_number)
                    ws.cell(row=current_row, column=4, value="Sin datos")
                    ws.cell(row=current_row, column=5, value="No hay registros")
                    ws.cell(row=current_row, column=6, value="")
                    ws.cell(row=current_row, column=7, value="--")
                    ws.cell(row=current_row, column=8, value="--")
                    ws.cell(row=current_row, column=9, value="00:00:00")
                    current_row += 1
            
            # Add summary row
            ws.cell(row=current_row, column=1, value="RESUMEN DEBUG")
            ws.cell(row=current_row, column=5, value=f"Total API calls: {self.api_calls}")
            ws.cell(row=current_row, column=6, value=f"Total time: {time.time() - self.start_time:.1f}s")
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            total_time = time.time() - self.start_time
            self.logger.info(f"=== REPORTE DEBUG COMPLETADO ===")
            self.logger.info(f"Total API calls: {self.api_calls}")
            self.logger.info(f"Total time: {total_time:.1f}s")
            self.logger.info(f"Empleados procesados: {len(employees)}")
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error generating debug report: {str(e)}")
            import traceback
            traceback.print_exc()
            return self._create_error_report(str(e))

    def _create_empty_report(self) -> bytes:
        """Create an empty report when no data is found"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Vacío"
        
        ws.cell(row=1, column=1, value="No se encontraron empleados")
        ws.cell(row=2, column=1, value="Verifique los filtros aplicados")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_error_report(self, error_message: str) -> bytes:
        """Create an error report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Error Debug"
        
        ws.cell(row=1, column=1, value="Error en generación debug")
        ws.cell(row=2, column=1, value=error_message)
        ws.cell(row=3, column=1, value=f"API calls realizadas: {self.api_calls}")
        ws.cell(row=4, column=1, value=f"Tiempo transcurrido: {time.time() - self.start_time:.1f}s")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_datetime(self, date_str: str) -> Optional[datetime]:
        """Parse datetime string from API"""
        if not date_str:
            return None
        
        try:
            return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        except ValueError:
            try:
                return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None

    def _get_employee_identification(self, employee: Dict) -> tuple:
        """Extract employee identification type and number"""
        identification_number = (
            employee.get('nid') or
            employee.get('identificationNumber') or
            employee.get('code') or
            str(employee.get('pin')) if employee.get('pin') else None or
            employee.get('id', 'Sin número')
        )
        
        identification_type = (
            employee.get('identityNumberType') or
            employee.get('identificationType') or
            'DNI'
        )
        
        return identification_type, str(identification_number)