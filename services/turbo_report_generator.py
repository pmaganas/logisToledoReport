"""
Generador de reportes ultra-rápido con optimizaciones avanzadas
"""
import asyncio
import openpyxl
import csv
import io
import gzip
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, List, Optional, Union, Callable, Any, Tuple
from datetime import datetime
import time
import json
from dataclasses import dataclass, asdict
from queue import Queue
import threading

from services.performance_optimizer import ParallelDataProcessor, AsyncAPIClient, PerformanceMetrics
from services.check_types_service import CheckTypesService
from utils.logging_config import get_logger
from config.settings import get_settings
from exceptions import ReportGenerationError, APIError
from models import SesameToken


@dataclass
class ReportConfig:
    """Configuración de generación de reporte"""
    from_date: Optional[str] = None
    to_date: Optional[str] = None
    employee_id: Optional[str] = None
    office_id: Optional[str] = None
    department_id: Optional[str] = None
    report_type: str = "by_employee"
    format_type: str = "xlsx"
    use_compression: bool = False
    enable_streaming: bool = True
    max_concurrent_requests: int = 15
    chunk_size: int = 1000


class StreamingExcelWriter:
    """Writer de Excel optimizado para streaming"""
    
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Reporte de Actividades"
        self.row_count = 1
        self.logger = get_logger(__name__)
        self._setup_styles()
    
    def _setup_styles(self):
        """Configurar estilos optimizados"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Estilos reutilizables
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill("solid", fgColor="366092")
        self.center_alignment = Alignment(horizontal="center")
        self.wrap_alignment = Alignment(wrap_text=True)
    
    def write_headers(self, headers: List[str]):
        """Escribir headers optimizado"""
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
        
        self.row_count = 2
    
    def write_batch(self, data_batch: List[List[Any]]):
        """Escribir batch de datos optimizado"""
        start_row = self.row_count
        
        # Escribir datos en batch para mejor performance
        for row_data in data_batch:
            for col, value in enumerate(row_data, 1):
                self.worksheet.cell(row=self.row_count, column=col, value=value)
            self.row_count += 1
        
        self.logger.debug(f"Batch escrito: filas {start_row}-{self.row_count-1}")
    
    def finalize(self) -> bytes:
        """Finalizar y retornar archivo Excel"""
        # Ajustar ancho de columnas
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Limitar ancho máximo para performance
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = io.BytesIO()
        self.workbook.save(output)
        return output.getvalue()


class StreamingCSVWriter:
    """Writer de CSV optimizado para streaming"""
    
    def __init__(self):
        self.output = io.StringIO()
        self.writer = csv.writer(self.output, quoting=csv.QUOTE_MINIMAL)
        self.row_count = 0
        self.logger = get_logger(__name__)
    
    def write_headers(self, headers: List[str]):
        """Escribir headers"""
        self.writer.writerow(headers)
        self.row_count += 1
    
    def write_batch(self, data_batch: List[List[Any]]):
        """Escribir batch de datos"""
        for row_data in data_batch:
            self.writer.writerow(row_data)
            self.row_count += 1
        
        self.logger.debug(f"CSV batch escrito: {len(data_batch)} filas")
    
    def finalize(self) -> bytes:
        """Finalizar y retornar archivo CSV"""
        csv_content = self.output.getvalue()
        # Añadir BOM para Excel
        return ('\ufeff' + csv_content).encode('utf-8')


class TurboReportGenerator:
    """Generador de reportes ultra-rápido"""
    
    def __init__(self):
        self.settings = get_settings()
        self.logger = get_logger(__name__)
        self.processor = ParallelDataProcessor(max_workers=20)
        self.check_types_service = CheckTypesService()
        self.metrics = PerformanceMetrics()
    
    async def generate_report_turbo(
        self,
        config: ReportConfig,
        progress_callback: Optional[Callable] = None
    ) -> bytes:
        """
        Generar reporte con optimizaciones turbo
        
        Args:
            config: Configuración del reporte
            progress_callback: Callback de progreso
            
        Returns:
            Datos del reporte en bytes
        """
        start_time = time.time()
        self.logger.info(f"🚀 Iniciando generación TURBO de reporte: {config.report_type}")
        
        try:
            # Paso 1: Verificar y cachear tipos de actividad
            await self._ensure_check_types_cached()
            
            # Paso 2: Obtener token y configurar cliente API
            token_info = SesameToken.get_active_token()
            if not token_info:
                raise ReportGenerationError("No hay token de API configurado")
            
            # Paso 3: Fetch de datos paralelo y asíncrono
            all_data = await self._fetch_data_parallel(token_info, config, progress_callback)
            
            # Paso 4: Procesamiento de datos en paralelo
            processed_data = await self._process_data_parallel(all_data, config)
            
            # Paso 5: Generación de archivo streaming
            report_bytes = await self._generate_file_streaming(processed_data, config)
            
            # Paso 6: Comprimir si está habilitado
            if config.use_compression:
                report_bytes = await self._compress_data(report_bytes)
            
            total_time = time.time() - start_time
            self.logger.info(
                f"🎯 Reporte TURBO completado: {len(processed_data)} registros en {total_time:.2f}s "
                f"({len(processed_data)/total_time:.1f} registros/s)"
            )
            
            return report_bytes
            
        except Exception as e:
            self.logger.error(f"❌ Error en generación TURBO: {str(e)}")
            raise ReportGenerationError(f"Error generando reporte turbo: {str(e)}", original_error=e)
    
    async def _ensure_check_types_cached(self):
        """Asegurar que tipos de actividad están en cache"""
        def fetch_check_types():
            return self.check_types_service.ensure_check_types_cached()
        
        cache_key = "check_types_cached"
        await asyncio.get_event_loop().run_in_executor(
            None, 
            self.processor.get_cached_or_fetch,
            cache_key,
            fetch_check_types
        )
    
    async def _fetch_data_parallel(
        self,
        token_info: Any,
        config: ReportConfig,
        progress_callback: Optional[Callable] = None
    ) -> List[Dict[str, Any]]:
        """Obtener datos de forma paralela y asíncrona"""
        base_url = f"https://api-{token_info.region}.sesametime.com"
        
        # Configurar parámetros de API
        params = {
            'employeeId': config.employee_id,
            'fromDate': config.from_date,
            'toDate': config.to_date,
            'officeId': config.office_id,
            'departmentId': config.department_id,
            'limit': self.settings.api.page_size
        }
        
        # Filtrar parámetros None
        params = {k: v for k, v in params.items() if v is not None}
        
        # Usar cliente asíncrono para máxima velocidad
        async with AsyncAPIClient(
            base_url=base_url,
            token=token_info.token,
            max_concurrent=config.max_concurrent_requests
        ) as client:
            data = await self.processor.fetch_all_pages_async(
                client,
                '/core/v3/time-tracking',
                params,
                progress_callback
            )
        
        self.logger.info(f"✅ Datos obtenidos: {len(data)} registros via requests paralelos")
        return data
    
    async def _process_data_parallel(
        self,
        data: List[Dict[str, Any]],
        config: ReportConfig
    ) -> List[Dict[str, Any]]:
        """Procesar datos en paralelo"""
        if not data:
            return []
        
        def process_entry(entry: Dict[str, Any]) -> Dict[str, Any]:
            """Procesar entrada individual"""
            try:
                # Obtener información básica
                employee = entry.get('employee', {})
                work_entry_in = entry.get('workEntryIn', {})
                work_entry_out = entry.get('workEntryOut', {})
                
                # Extraer datos del empleado
                employee_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                employee_code = employee.get('code', '')
                
                # Procesar fechas y horas
                date_str = work_entry_in.get('date', '').split('T')[0] if work_entry_in.get('date') else ''
                
                start_time = ''
                end_time = ''
                if work_entry_in.get('date'):
                    start_dt = datetime.fromisoformat(work_entry_in['date'].replace('Z', '+00:00'))
                    start_time = start_dt.strftime('%H:%M:%S')
                
                if work_entry_out and work_entry_out.get('date'):
                    end_dt = datetime.fromisoformat(work_entry_out['date'].replace('Z', '+00:00'))
                    end_time = end_dt.strftime('%H:%M:%S')
                
                # Calcular duración
                duration_seconds = entry.get('workedSeconds', 0)
                duration = self._format_duration(duration_seconds)
                
                # Obtener tipo de actividad
                check_type_id = work_entry_in.get('checkTypeId', '')
                activity_name = self.check_types_service.get_activity_name(check_type_id)
                
                return {
                    'date': date_str,
                    'employee_name': employee_name,
                    'employee_code': employee_code,
                    'start_time': start_time,
                    'end_time': end_time,
                    'duration': duration,
                    'activity_name': activity_name,
                    'worked_seconds': duration_seconds
                }
                
            except Exception as e:
                self.logger.warning(f"Error procesando entrada: {str(e)}")
                return {
                    'date': '', 'employee_name': '', 'employee_code': '',
                    'start_time': '', 'end_time': '', 'duration': '00:00:00',
                    'activity_name': 'Error', 'worked_seconds': 0
                }
        
        # Procesar en paralelo usando ThreadPoolExecutor
        processed_data = await asyncio.get_event_loop().run_in_executor(
            None,
            self.processor.process_data_parallel,
            data,
            process_entry,
            config.chunk_size
        )
        
        # Ordenar por fecha y empleado
        processed_data.sort(key=lambda x: (x.get('date', ''), x.get('employee_name', '')))
        
        self.logger.info(f"✅ Datos procesados: {len(processed_data)} registros")
        return processed_data
    
    async def _generate_file_streaming(
        self,
        data: List[Dict[str, Any]],
        config: ReportConfig
    ) -> bytes:
        """Generar archivo usando streaming para mejor performance"""
        
        def generate_file():
            if config.format_type == 'csv':
                return self._generate_csv_streaming(data)
            else:
                return self._generate_excel_streaming(data)
        
        # Ejecutar generación en thread separado para no bloquear
        file_bytes = await asyncio.get_event_loop().run_in_executor(None, generate_file)
        
        self.logger.info(f"✅ Archivo generado: {len(file_bytes)} bytes ({config.format_type.upper()})")
        return file_bytes
    
    def _generate_excel_streaming(self, data: List[Dict[str, Any]]) -> bytes:
        """Generar Excel usando streaming writer"""
        writer = StreamingExcelWriter()
        
        # Headers
        headers = [
            'Fecha', 'Empleado', 'Código', 'Hora Entrada',
            'Hora Salida', 'Duración', 'Actividad'
        ]
        writer.write_headers(headers)
        
        # Procesar datos en batches para mejor memoria
        batch_size = 500
        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            batch_rows = []
            
            for entry in batch:
                row = [
                    entry.get('date', ''),
                    entry.get('employee_name', ''),
                    entry.get('employee_code', ''),
                    entry.get('start_time', ''),
                    entry.get('end_time', ''),
                    entry.get('duration', ''),
                    entry.get('activity_name', '')
                ]
                batch_rows.append(row)
            
            writer.write_batch(batch_rows)
        
        return writer.finalize()
    
    def _generate_csv_streaming(self, data: List[Dict[str, Any]]) -> bytes:
        """Generar CSV usando streaming writer"""
        writer = StreamingCSVWriter()
        
        # Headers
        headers = [
            'Fecha', 'Empleado', 'Código', 'Hora Entrada',
            'Hora Salida', 'Duración', 'Actividad'
        ]
        writer.write_headers(headers)
        
        # Procesar datos en batches
        batch_size = 1000  # CSV puede manejar batches más grandes
        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            batch_rows = []
            
            for entry in batch:
                row = [
                    entry.get('date', ''),
                    entry.get('employee_name', ''),
                    entry.get('employee_code', ''),
                    entry.get('start_time', ''),
                    entry.get('end_time', ''),
                    entry.get('duration', ''),
                    entry.get('activity_name', '')
                ]
                batch_rows.append(row)
            
            writer.write_batch(batch_rows)
        
        return writer.finalize()
    
    async def _compress_data(self, data: bytes) -> bytes:
        """Comprimir datos usando gzip"""
        def compress():
            return gzip.compress(data, compresslevel=6)
        
        compressed = await asyncio.get_event_loop().run_in_executor(None, compress)
        compression_ratio = len(compressed) / len(data)
        
        self.logger.info(f"📦 Compresión aplicada: {compression_ratio:.2%} del tamaño original")
        return compressed
    
    def _format_duration(self, seconds: int) -> str:
        """Formatear duración en HH:MM:SS"""
        if not seconds:
            return "00:00:00"
        
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    
    def get_performance_metrics(self) -> Dict[str, Any]:
        """Obtener métricas de performance"""
        return {
            'turbo_metrics': self.metrics.to_dict(),
            'processor_metrics': self.processor.get_performance_metrics()
        }


# Función de conveniencia para usar el generador turbo
async def generate_turbo_report(
    config: ReportConfig,
    progress_callback: Optional[Callable] = None
) -> bytes:
    """
    Generar reporte usando optimizaciones turbo
    
    Args:
        config: Configuración del reporte
        progress_callback: Callback de progreso opcional
        
    Returns:
        Datos del reporte en bytes
    """
    generator = TurboReportGenerator()
    return await generator.generate_report_turbo(config, progress_callback)