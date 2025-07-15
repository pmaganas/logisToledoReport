import openpyxl
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from services.sesame_api import SesameAPI

class BasicReportGenerator:
    def __init__(self):
        self.sesame_api = SesameAPI()
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)

    def generate_basic_report(self, from_date: str = None, to_date: str = None, 
                             employee_id: str = None, office_id: str = None, 
                             department_id: str = None, report_type: str = "by_employee") -> Optional[bytes]:
        """Generate a basic XLSX report with minimal API calls to avoid SSL issues"""
        
        try:
            self.logger.info(f"PASO 1/4: Iniciando generación de reporte básico - Tipo: {report_type}")
            
            # Step 1: Get specific employee or limit to first few employees
            if employee_id:
                employee_response = self.sesame_api.get_employee_details(employee_id)
                employees = [employee_response.get('data')] if employee_response else []
            else:
                # Get only first page of employees to avoid long connections
                employees_response = self.sesame_api.get_employees(page=1, per_page=10)
                employees = employees_response.get('data', []) if employees_response else []
                
                # Apply filters
                if office_id or department_id:
                    filtered_employees = []
                    for employee in employees:
                        include_employee = True
                        
                        if office_id:
                            employee_office_id = employee.get('office', {}).get('id') if employee.get('office') else None
                            if employee_office_id != office_id:
                                include_employee = False
                        
                        if department_id and include_employee:
                            employee_department_id = employee.get('department', {}).get('id') if employee.get('department') else None
                            if employee_department_id != department_id:
                                include_employee = False
                        
                        if include_employee:
                            filtered_employees.append(employee)
                    
                    employees = filtered_employees
            
            if not employees:
                self.logger.warning("No employees found")
                return self._create_empty_report()
            
            # Limit to first 5 employees to avoid timeouts
            employees = employees[:5]
            self.logger.info(f"PASO 2/4: Procesando {len(employees)} empleados")
            
            # Step 2: Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Básico"
            
            # Headers
            headers = ["Empleado", "Tipo ID", "Número ID", "Fecha", "Actividad", "Grupo", "Entrada", "Salida", "Tiempo Registrado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            current_row = 2
            
            self.logger.info("PASO 3/4: Cargando datos con llamadas limitadas...")
            
            # Step 3: Process each employee with minimal API calls
            for i, employee in enumerate(employees, 1):
                emp_id = employee.get('id')
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                identity_type, identity_number = self._get_employee_identification(employee)
                
                self.logger.info(f"PASO 3/4: Procesando empleado {i}/{len(employees)}: {emp_name}")
                
                try:
                    # Get only first page of time entries to avoid long connections
                    time_response = self.sesame_api.get_work_entries(
                        employee_id=emp_id,
                        from_date=from_date,
                        to_date=to_date,
                        page=1,
                        limit=20  # Limit to 20 entries per employee
                    )
                    
                    if time_response and time_response.get('data'):
                        time_entries = time_response['data']
                        
                        # Group entries by date
                        entries_by_date = {}
                        for entry in time_entries:
                            work_in = entry.get('workEntryIn', {})
                            if work_in and work_in.get('date'):
                                try:
                                    parsed_date = self._parse_datetime(work_in.get('date'))
                                    if parsed_date:
                                        date_key = parsed_date.strftime('%Y-%m-%d')
                                        if date_key not in entries_by_date:
                                            entries_by_date[date_key] = []
                                        entries_by_date[date_key].append(entry)
                                except Exception:
                                    continue
                        
                        # Process each date
                        for date_key in sorted(entries_by_date.keys()):
                            date_entries = entries_by_date[date_key]
                            daily_total_seconds = 0
                            
                            # Add individual entries for this date
                            for entry in date_entries:
                                work_in = entry.get('workEntryIn', {})
                                work_out = entry.get('workEntryOut', {})
                                
                                start_time = self._parse_datetime(work_in.get('date'))
                                end_time = self._parse_datetime(work_out.get('date'))
                                
                                # Calculate duration
                                duration_str = "00:00:00"
                                entry_seconds = 0
                                if start_time and end_time:
                                    duration = end_time - start_time
                                    entry_seconds = duration.total_seconds()
                                    daily_total_seconds += entry_seconds
                                    hours = int(entry_seconds // 3600)
                                    minutes = int((entry_seconds % 3600) // 60)
                                    seconds = int(entry_seconds % 60)
                                    duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                                
                                # Get activity name
                                activity_name = entry.get('workEntryType', 'Trabajo')
                                
                                # Add row for this entry
                                ws.cell(row=current_row, column=1, value=emp_name)
                                ws.cell(row=current_row, column=2, value=identity_type)
                                ws.cell(row=current_row, column=3, value=identity_number)
                                ws.cell(row=current_row, column=4, value=date_key)
                                ws.cell(row=current_row, column=5, value=activity_name)
                                ws.cell(row=current_row, column=6, value="")  # Group column empty
                                ws.cell(row=current_row, column=7, value=start_time.strftime("%H:%M:%S") if start_time else "N/A")
                                ws.cell(row=current_row, column=8, value=end_time.strftime("%H:%M:%S") if end_time else "N/A")
                                ws.cell(row=current_row, column=9, value=duration_str)
                                current_row += 1
                            
                            # Add total row for this date
                            daily_total_hours = int(daily_total_seconds // 3600)
                            daily_total_minutes = int((daily_total_seconds % 3600) // 60)
                            daily_total_secs = int(daily_total_seconds % 60)
                            daily_total_str = f"{daily_total_hours:02d}:{daily_total_minutes:02d}:{daily_total_secs:02d}"
                            
                            # Add TOTAL row
                            ws.cell(row=current_row, column=1, value="TOTAL")
                            ws.cell(row=current_row, column=2, value="")
                            ws.cell(row=current_row, column=3, value="")
                            ws.cell(row=current_row, column=4, value="")
                            ws.cell(row=current_row, column=5, value="Total del día")
                            ws.cell(row=current_row, column=6, value="")
                            ws.cell(row=current_row, column=7, value="")
                            ws.cell(row=current_row, column=8, value="")
                            ws.cell(row=current_row, column=9, value=daily_total_str)
                            
                            # Style the TOTAL row
                            for col in range(1, 10):
                                cell = ws.cell(row=current_row, column=col)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                            
                            current_row += 1
                    
                except Exception as e:
                    self.logger.error(f"Error processing employee {emp_name}: {str(e)}")
                    # Add error row
                    ws.cell(row=current_row, column=1, value=emp_name)
                    ws.cell(row=current_row, column=2, value=identity_type)
                    ws.cell(row=current_row, column=3, value=identity_number)
                    ws.cell(row=current_row, column=4, value="Error")
                    ws.cell(row=current_row, column=5, value="Error al cargar datos")
                    ws.cell(row=current_row, column=6, value="")
                    ws.cell(row=current_row, column=7, value="Error")
                    ws.cell(row=current_row, column=8, value="Error")
                    ws.cell(row=current_row, column=9, value="00:00:00")
                    current_row += 1
            
            self.logger.info("PASO 4/4: Generando archivo Excel...")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            self.logger.info("✓ Reporte básico generado exitosamente")
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error generating basic report: {str(e)}")
            return self._create_error_report(str(e))

    def _create_empty_report(self) -> bytes:
        """Create an empty report when no data is found"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Vacío"
        
        ws.cell(row=1, column=1, value="No se encontraron empleados")
        ws.cell(row=2, column=1, value="Verifique los filtros aplicados")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_error_report(self, error_message: str) -> bytes:
        """Create an error report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Error"
        
        ws.cell(row=1, column=1, value="Error al generar el reporte")
        ws.cell(row=2, column=1, value=error_message)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_datetime(self, date_str: str) -> Optional[datetime]:
        """Parse datetime string from API"""
        if not date_str:
            return None
        
        try:
            return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        except ValueError:
            try:
                return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None

    def _get_employee_identification(self, employee: Dict) -> tuple:
        """Extract employee identification type and number"""
        identification_number = (
            employee.get('nid') or
            employee.get('identificationNumber') or
            employee.get('code') or
            str(employee.get('pin')) if employee.get('pin') else None or
            employee.get('id', 'Sin número')
        )
        
        identification_type = (
            employee.get('identityNumberType') or
            employee.get('identificationType') or
            'DNI'
        )
        
        return identification_type, str(identification_number)