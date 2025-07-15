import openpyxl
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from services.sesame_api import SesameAPI

class UltraBasicReportGenerator:
    def __init__(self):
        self.sesame_api = SesameAPI()
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)

    def generate_ultra_basic_report(self, from_date: str = None, to_date: str = None, 
                                   employee_id: str = None, office_id: str = None, 
                                   department_id: str = None, report_type: str = "by_employee") -> Optional[bytes]:
        """Generate an ultra-basic XLSX report with absolute minimum API calls"""
        
        try:
            self.logger.info("PASO 1/3: Iniciando generación ultra-básica de reporte")
            
            # Step 1: Get ONE page of employees only
            if employee_id:
                # Get specific employee details
                employee_response = self.sesame_api.get_employee_details(employee_id)
                employees = [employee_response.get('data')] if employee_response else []
            else:
                # Get only FIRST page with minimal employees
                employees_response = self.sesame_api.get_employees(page=1, per_page=5)
                employees = employees_response.get('data', [])[:3] if employees_response else []
            
            if not employees:
                self.logger.warning("No employees found")
                return self._create_empty_report()
            
            self.logger.info(f"PASO 2/3: Procesando {len(employees)} empleados con datos mínimos")
            
            # Step 2: Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Ultra-Básico"
            
            # Headers
            headers = ["Empleado", "Tipo ID", "Número ID", "Estado", "Mensaje"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            current_row = 2
            
            # Step 3: Process employees with MINIMAL data
            for employee in employees:
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                identity_type, identity_number = self._get_employee_identification(employee)
                
                # Just show employee info without time data to avoid SSL issues
                ws.cell(row=current_row, column=1, value=emp_name)
                ws.cell(row=current_row, column=2, value=identity_type)
                ws.cell(row=current_row, column=3, value=identity_number)
                ws.cell(row=current_row, column=4, value="Activo")
                ws.cell(row=current_row, column=5, value="Empleado encontrado - Datos de tiempo no disponibles por problemas SSL")
                current_row += 1
            
            # Add info row
            ws.cell(row=current_row, column=1, value="INFORMACIÓN")
            ws.cell(row=current_row, column=2, value="")
            ws.cell(row=current_row, column=3, value="")
            ws.cell(row=current_row, column=4, value="SISTEMA")
            ws.cell(row=current_row, column=5, value="Este reporte usa datos mínimos para evitar errores de conexión SSL")
            
            self.logger.info("PASO 3/3: Generando archivo Excel ultra-básico...")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            self.logger.info("✓ Reporte ultra-básico generado exitosamente")
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error generating ultra-basic report: {str(e)}")
            return self._create_error_report(str(e))

    def _create_empty_report(self) -> bytes:
        """Create an empty report when no data is found"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Vacío"
        
        ws.cell(row=1, column=1, value="No se encontraron empleados")
        ws.cell(row=2, column=1, value="Verifique los filtros aplicados")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_error_report(self, error_message: str) -> bytes:
        """Create an error report with diagnostic information"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diagnóstico de Error"
        
        ws.cell(row=1, column=1, value="DIAGNÓSTICO DE ERROR SSL")
        ws.cell(row=2, column=1, value="Error específico:")
        ws.cell(row=3, column=1, value=error_message)
        ws.cell(row=4, column=1, value="")
        ws.cell(row=5, column=1, value="POSIBLES SOLUCIONES:")
        ws.cell(row=6, column=1, value="1. El servidor API puede estar sobrecargado")
        ws.cell(row=7, column=1, value="2. Conexión SSL inestable")
        ws.cell(row=8, column=1, value="3. Timeout de red")
        ws.cell(row=9, column=1, value="4. Límite de rate limiting alcanzado")
        ws.cell(row=10, column=1, value="")
        ws.cell(row=11, column=1, value="RECOMENDACIONES:")
        ws.cell(row=12, column=1, value="- Intentar más tarde")
        ws.cell(row=13, column=1, value="- Filtrar por empleado específico")
        ws.cell(row=14, column=1, value="- Usar rango de fechas más pequeño")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _get_employee_identification(self, employee: Dict) -> tuple:
        """Extract employee identification type and number"""
        identification_number = (
            employee.get('nid') or
            employee.get('identificationNumber') or
            employee.get('code') or
            str(employee.get('pin')) if employee.get('pin') else None or
            employee.get('id', 'Sin número')
        )
        
        identification_type = (
            employee.get('identityNumberType') or
            employee.get('identificationType') or
            'DNI'
        )
        
        return identification_type, str(identification_number)