import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import logging
import io
import os
from .sesame_api import SesameAPI

class ReportGenerator:
    def __init__(self):
        self.sesame_api = SesameAPI()
        self.logger = logging.getLogger(__name__)

    def _parse_datetime(self, date_str: str) -> Optional[datetime]:
        """Parse datetime string from API"""
        if not date_str:
            return None
            
        try:
            # Handle different datetime formats
            formats = [
                "%Y-%m-%dT%H:%M:%S%z",
                "%Y-%m-%dT%H:%M:%S.%f%z",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d %H:%M:%S"
            ]
            
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
                    
            return None
        except Exception as e:
            self.logger.error(f"Error parsing datetime {date_str}: {str(e)}")
            return None

    def _calculate_duration(self, start_time: datetime, end_time: datetime) -> timedelta:
        """Calculate duration between two datetime objects"""
        if not start_time or not end_time:
            return timedelta(0)
        return end_time - start_time

    def _format_duration(self, duration: timedelta) -> str:
        """Format duration as HH:MM:SS"""
        total_seconds = int(duration.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def _merge_breakfast_breaks(self, activities: List[Dict], breaks: List[Dict]) -> List[Dict]:
        """Merge breakfast break time with adjacent activities"""
        merged_activities = []
        
        # Sort activities and breaks by start time
        activities.sort(key=lambda x: self._parse_datetime(x.get('startTime', '')) or datetime.min)
        breaks.sort(key=lambda x: self._parse_datetime(x.get('startTime', '')) or datetime.min)
        
        for activity in activities:
            activity_start = self._parse_datetime(activity.get('startTime'))
            activity_end = self._parse_datetime(activity.get('endTime'))
            
            if not activity_start or not activity_end:
                continue
                
            # Find breakfast breaks that occur during or adjacent to this activity
            breakfast_breaks = []
            for break_entry in breaks:
                break_type = break_entry.get('type', '').lower()
                if 'breakfast' in break_type or 'desayuno' in break_type:
                    break_start = self._parse_datetime(break_entry.get('startTime'))
                    break_end = self._parse_datetime(break_entry.get('endTime'))
                    
                    if break_start and break_end:
                        # Check if break is adjacent to or overlaps with activity
                        if (break_start <= activity_end and break_end >= activity_start):
                            breakfast_breaks.append(break_entry)
            
            # Merge breakfast breaks with activity
            merged_activity = activity.copy()
            
            if breakfast_breaks:
                # Calculate total breakfast time
                total_breakfast_time = timedelta(0)
                for break_entry in breakfast_breaks:
                    break_start = self._parse_datetime(break_entry.get('startTime'))
                    break_end = self._parse_datetime(break_entry.get('endTime'))
                    if break_start and break_end:
                        total_breakfast_time += self._calculate_duration(break_start, break_end)
                
                # Add breakfast time to activity duration
                original_duration = self._calculate_duration(activity_start, activity_end)
                merged_activity['totalTime'] = original_duration + total_breakfast_time
                
                # Update activity notes to indicate breakfast break inclusion
                notes = merged_activity.get('notes', '')
                if notes:
                    notes += f" (Incluye {self._format_duration(total_breakfast_time)} de descanso de desayuno)"
                else:
                    notes = f"Incluye {self._format_duration(total_breakfast_time)} de descanso de desayuno"
                merged_activity['notes'] = notes
            
            merged_activities.append(merged_activity)
        
        return merged_activities

    def _get_employee_identification(self, employee: Dict) -> tuple:
        """Extract employee identification type and number"""
        identity_type = employee.get('identityNumberType', 'N/A')
        identity_number = employee.get('nid', employee.get('ssn', 'N/A'))
        
        # Map identity types to Spanish
        type_mapping = {
            'dni': 'DNI',
            'nie': 'NIE',
            'rut': 'RUT',
            'other': 'Otro'
        }
        
        return type_mapping.get(identity_type, identity_type), identity_number

    def generate_report(self, from_date: str = None, to_date: str = None, 
                       employee_id: str = None, office_id: str = None, department_id: str = None, 
                       company_id: str = None, report_type: str = "by_employee") -> Optional[bytes]:
        """Generate XLSX report with employee activities grouped by date and employee with totals"""
        try:
            # Get token info to get company ID if not provided
            if not company_id:
                token_info = self.sesame_api.get_token_info()
                if token_info and token_info.get('data', {}).get('company', {}).get('id'):
                    company_id = token_info['data']['company']['id']
                else:
                    self.logger.error("Could not get company ID from token info")
                    return None

            # Get employees based on filters
            if employee_id:
                # Get specific employee
                employee_data = self.sesame_api.get_employee_details(employee_id)
                employees = [employee_data['data']] if employee_data else []
            else:
                # Get all employees and filter by office and department
                all_employees = self.sesame_api.get_all_employees_data(company_id)
                employees = []
                
                # Filter by office and department
                for employee in all_employees:
                    include_employee = True
                    
                    # Filter by office_id
                    if office_id:
                        employee_office_id = employee.get('office', {}).get('id') if employee.get('office') else None
                        if employee_office_id != office_id:
                            include_employee = False
                    
                    # Filter by department_id
                    if department_id and include_employee:
                        employee_department_id = employee.get('department', {}).get('id') if employee.get('department') else None
                        if employee_department_id != department_id:
                            include_employee = False
                    
                    if include_employee:
                        employees.append(employee)

            if not employees:
                self.logger.error("No employees found")
                return None

            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Set worksheet title based on report type
            if report_type == "by_employee":
                ws.title = "Por Empleado"
            elif report_type == "by_type":
                ws.title = "Por Tipo de Fichaje"
            elif report_type == "by_group":
                ws.title = "Por Grupos"
            else:
                ws.title = "Reporte de Actividades"

            # Define headers
            headers = [
                "Empleado",
                "Tipo de Identificación",
                "Nº de Identificación",
                "Fecha",
                "Actividad",
                "Grupo",
                "Entrada",
                "Salida",
                "Tiempo registrado"
            ]

            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Generate report based on type
            if report_type == "by_employee":
                return self._generate_by_employee_report(wb, ws, employees, from_date, to_date, company_id)
            elif report_type == "by_type":
                return self._generate_by_type_report(wb, ws, employees, from_date, to_date, company_id)
            elif report_type == "by_group":
                return self._generate_by_group_report(wb, ws, employees, from_date, to_date, company_id)
            else:
                return self._generate_by_employee_report(wb, ws, employees, from_date, to_date, company_id)

        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}", exc_info=True)
            return None

    def _generate_by_employee_report(self, wb, ws, employees, from_date, to_date, company_id):
        """Generate report grouped by employee"""
        try:
            # Collect and organize data by employee and date
            employee_date_data = {}
            row = 2
            
            for employee in employees:
                emp_id = employee.get('id')
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                identity_type, identity_number = self._get_employee_identification(employee)

                # Get all work entries for this employee
                time_entries = self.sesame_api.get_all_time_tracking_data(
                    employee_id=emp_id,
                    from_date=from_date,
                    to_date=to_date
                )

                # Process work entries grouped by date
                date_entries = {}
                for entry in time_entries:
                    work_in = entry.get('workEntryIn', {})
                    work_out = entry.get('workEntryOut', {})
                    
                    start_time = self._parse_datetime(work_in.get('date')) if work_in else None
                    end_time = self._parse_datetime(work_out.get('date')) if work_out else None
                    
                    if not start_time:
                        continue

                    date_key = start_time.strftime("%Y-%m-%d")
                    
                    if date_key not in date_entries:
                        date_entries[date_key] = []
                    
                    # Calculate registered time
                    if start_time and end_time:
                        duration = self._calculate_duration(start_time, end_time)
                    elif entry.get('workedSeconds'):
                        duration = timedelta(seconds=entry['workedSeconds'])
                    else:
                        duration = timedelta(0)

                    date_entries[date_key].append({
                        'employee': emp_name,
                        'identity_type': identity_type,
                        'identity_number': identity_number,
                        'date': date_key,
                        'activity': entry.get('workEntryType', 'Trabajo'),
                        'group': 'Actividad Principal',
                        'start_time': start_time.strftime("%H:%M:%S") if start_time else "N/A",
                        'end_time': end_time.strftime("%H:%M:%S") if end_time else "N/A",
                        'duration': duration,
                        'registered_time': self._format_duration(duration)
                    })

                # Add data to worksheet grouped by employee and date
                for date_key in sorted(date_entries.keys()):
                    entries = date_entries[date_key]
                    
                    # Add individual entries
                    for entry in entries:
                        ws.cell(row=row, column=1, value=entry['employee'])
                        ws.cell(row=row, column=2, value=entry['identity_type'])
                        ws.cell(row=row, column=3, value=entry['identity_number'])
                        ws.cell(row=row, column=4, value=entry['date'])
                        ws.cell(row=row, column=5, value=entry['activity'])
                        ws.cell(row=row, column=6, value=entry['group'])
                        ws.cell(row=row, column=7, value=entry['start_time'])
                        ws.cell(row=row, column=8, value=entry['end_time'])
                        ws.cell(row=row, column=9, value=entry['registered_time'])
                        row += 1

                    # Add total row for this employee and date
                    if entries:
                        total_duration = sum([entry['duration'] for entry in entries], timedelta(0))
                        total_entries = len(entries)
                        
                        # Total row styling
                        total_row = row
                        ws.cell(row=total_row, column=1, value=f"TOTAL - {emp_name}")
                        ws.cell(row=total_row, column=2, value="")
                        ws.cell(row=total_row, column=3, value="")
                        ws.cell(row=total_row, column=4, value=date_key)
                        ws.cell(row=total_row, column=5, value=str(total_entries))
                        ws.cell(row=total_row, column=6, value="")
                        ws.cell(row=total_row, column=7, value="")
                        ws.cell(row=total_row, column=8, value="")
                        ws.cell(row=total_row, column=9, value=self._format_duration(total_duration))
                        
                        # Style total row
                        for col in range(1, 10):
                            cell = ws.cell(row=total_row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                        
                        row += 1

            # Add a message if no data found
            if row == 2:
                ws.cell(row=2, column=1, value="No se encontraron datos para el período seleccionado")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()

        except Exception as e:
            self.logger.error(f"Error generating by employee report: {str(e)}", exc_info=True)
            return None

    def _generate_by_type_report(self, wb, ws, employees, from_date, to_date, company_id):
        """Generate report grouped by type of fichaje"""
        try:
            # Collect all work entries for all employees
            all_entries = []
            for employee in employees:
                emp_id = employee.get('id')
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                identity_type, identity_number = self._get_employee_identification(employee)

                # Get all work entries for this employee
                work_entries = self.sesame_api.get_all_time_tracking_data(
                    employee_id=emp_id,
                    from_date=from_date,
                    to_date=to_date
                )
                
                for entry in work_entries:
                    work_in = entry.get('workEntryIn', {})
                    work_out = entry.get('workEntryOut', {})
                    
                    start_time = self._parse_datetime(work_in.get('date')) if work_in else None
                    end_time = self._parse_datetime(work_out.get('date')) if work_out else None
                    
                    if not start_time:
                        continue

                    # Calculate registered time
                    if start_time and end_time:
                        duration = self._calculate_duration(start_time, end_time)
                    elif entry.get('workedSeconds'):
                        duration = timedelta(seconds=entry['workedSeconds'])
                    else:
                        duration = timedelta(0)

                    fichaje_type = entry.get('workEntryType', 'Trabajo')
                    
                    all_entries.append({
                        'employee': emp_name,
                        'identity_type': identity_type,
                        'identity_number': identity_number,
                        'date': start_time.strftime("%Y-%m-%d"),
                        'activity': fichaje_type,
                        'group': 'Actividad Principal',
                        'start_time': start_time.strftime("%H:%M:%S") if start_time else "N/A",
                        'end_time': end_time.strftime("%H:%M:%S") if end_time else "N/A",
                        'duration': duration,
                        'registered_time': self._format_duration(duration),
                        'fichaje_type': fichaje_type
                    })

            # Group by fichaje type
            type_groups = {}
            for entry in all_entries:
                fichaje_type = entry['fichaje_type']
                if fichaje_type not in type_groups:
                    type_groups[fichaje_type] = []
                type_groups[fichaje_type].append(entry)

            # Add data to worksheet grouped by fichaje type
            row = 2
            for fichaje_type in sorted(type_groups.keys()):
                entries = type_groups[fichaje_type]
                
                # Add individual entries
                for entry in entries:
                    ws.cell(row=row, column=1, value=entry['employee'])
                    ws.cell(row=row, column=2, value=entry['identity_type'])
                    ws.cell(row=row, column=3, value=entry['identity_number'])
                    ws.cell(row=row, column=4, value=entry['date'])
                    ws.cell(row=row, column=5, value=entry['activity'])
                    ws.cell(row=row, column=6, value=entry['group'])
                    ws.cell(row=row, column=7, value=entry['start_time'])
                    ws.cell(row=row, column=8, value=entry['end_time'])
                    ws.cell(row=row, column=9, value=entry['registered_time'])
                    row += 1

                # Add total row for this fichaje type
                if entries:
                    total_duration = sum([entry['duration'] for entry in entries], timedelta(0))
                    total_entries = len(entries)
                    
                    # Total row
                    total_row = row
                    ws.cell(row=total_row, column=1, value=f"TOTAL - {fichaje_type}")
                    ws.cell(row=total_row, column=2, value="")
                    ws.cell(row=total_row, column=3, value="")
                    ws.cell(row=total_row, column=4, value="")
                    ws.cell(row=total_row, column=5, value=str(total_entries))
                    ws.cell(row=total_row, column=6, value="")
                    ws.cell(row=total_row, column=7, value="")
                    ws.cell(row=total_row, column=8, value="")
                    ws.cell(row=total_row, column=9, value=self._format_duration(total_duration))
                    
                    # Style total row
                    for col in range(1, 10):
                        cell = ws.cell(row=total_row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    
                    row += 1

            # Add a message if no data found
            if row == 2:
                ws.cell(row=2, column=1, value="No se encontraron datos para el período seleccionado")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()

        except Exception as e:
            self.logger.error(f"Error generating by type report: {str(e)}", exc_info=True)
            return None

    def _generate_by_group_report(self, wb, ws, employees, from_date, to_date, company_id):
        """Generate report grouped by groups"""
        try:
            # Collect all work entries for all employees
            all_entries = []
            for employee in employees:
                emp_id = employee.get('id')
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()
                identity_type, identity_number = self._get_employee_identification(employee)

                # Get all work entries for this employee
                work_entries = self.sesame_api.get_all_time_tracking_data(
                    employee_id=emp_id,
                    from_date=from_date,
                    to_date=to_date
                )
                
                for entry in work_entries:
                    work_in = entry.get('workEntryIn', {})
                    work_out = entry.get('workEntryOut', {})
                    
                    start_time = self._parse_datetime(work_in.get('date')) if work_in else None
                    end_time = self._parse_datetime(work_out.get('date')) if work_out else None
                    
                    if not start_time:
                        continue

                    # Calculate registered time
                    if start_time and end_time:
                        duration = self._calculate_duration(start_time, end_time)
                    elif entry.get('workedSeconds'):
                        duration = timedelta(seconds=entry['workedSeconds'])
                    else:
                        duration = timedelta(0)

                    # Try to get group from activity or use default
                    group = 'Actividad Principal'
                    if entry.get('activity', {}).get('group', {}).get('name'):
                        group = entry['activity']['group']['name']
                    
                    all_entries.append({
                        'employee': emp_name,
                        'identity_type': identity_type,
                        'identity_number': identity_number,
                        'date': start_time.strftime("%Y-%m-%d"),
                        'activity': entry.get('workEntryType', 'Trabajo'),
                        'group': group,
                        'start_time': start_time.strftime("%H:%M:%S") if start_time else "N/A",
                        'end_time': end_time.strftime("%H:%M:%S") if end_time else "N/A",
                        'duration': duration,
                        'registered_time': self._format_duration(duration),
                        'group_key': group
                    })

            # Group by group
            group_groups = {}
            for entry in all_entries:
                group_key = entry['group_key']
                if group_key not in group_groups:
                    group_groups[group_key] = []
                group_groups[group_key].append(entry)

            # Add data to worksheet grouped by group
            row = 2
            for group_key in sorted(group_groups.keys()):
                entries = group_groups[group_key]
                
                # Add individual entries
                for entry in entries:
                    ws.cell(row=row, column=1, value=entry['employee'])
                    ws.cell(row=row, column=2, value=entry['identity_type'])
                    ws.cell(row=row, column=3, value=entry['identity_number'])
                    ws.cell(row=row, column=4, value=entry['date'])
                    ws.cell(row=row, column=5, value=entry['activity'])
                    ws.cell(row=row, column=6, value=entry['group'])
                    ws.cell(row=row, column=7, value=entry['start_time'])
                    ws.cell(row=row, column=8, value=entry['end_time'])
                    ws.cell(row=row, column=9, value=entry['registered_time'])
                    row += 1

                # Add total row for this group
                if entries:
                    total_duration = sum([entry['duration'] for entry in entries], timedelta(0))
                    total_entries = len(entries)
                    
                    # Total row
                    total_row = row
                    ws.cell(row=total_row, column=1, value=f"TOTAL - {group_key}")
                    ws.cell(row=total_row, column=2, value="")
                    ws.cell(row=total_row, column=3, value="")
                    ws.cell(row=total_row, column=4, value="")
                    ws.cell(row=total_row, column=5, value=str(total_entries))
                    ws.cell(row=total_row, column=6, value="")
                    ws.cell(row=total_row, column=7, value="")
                    ws.cell(row=total_row, column=8, value="")
                    ws.cell(row=total_row, column=9, value=self._format_duration(total_duration))
                    
                    # Style total row
                    for col in range(1, 10):
                        cell = ws.cell(row=total_row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    
                    row += 1

            # Add a message if no data found
            if row == 2:
                ws.cell(row=2, column=1, value="No se encontraron datos para el período seleccionado")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()

        except Exception as e:
            self.logger.error(f"Error generating by group report: {str(e)}", exc_info=True)
            return None

    def test_connection(self) -> Dict:
        """Test API connection and return status"""
        try:
            token_info = self.sesame_api.get_token_info()
            if token_info:
                return {
                    "status": "success",
                    "message": "Connection successful",
                    "company": token_info.get('data', {}).get('company', {}).get('name', 'Unknown')
                }
            else:
                return {
                    "status": "error",
                    "message": "Failed to connect to Sesame API"
                }
        except Exception as e:
            return {
                "status": "error",
                "message": f"Connection test failed: {str(e)}"
            }
