import openpyxl
import logging
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from services.sesame_api import SesameAPI


class OptimizedReportGenerator:

    def __init__(self):
        self.sesame_api = SesameAPI()
        self.logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.INFO)
        self.api_calls = 0
        self.start_time = time.time()

    def log_api_call(self, endpoint: str, employee_name: str = None):
        """Log each API call with timing"""
        self.api_calls += 1
        elapsed = time.time() - self.start_time
        context = f" for {employee_name}" if employee_name else ""
        self.logger.info(
            f"API CALL #{self.api_calls}: {endpoint}{context} [Total time: {elapsed:.1f}s]"
        )

    def get_data_metrics(self,
                         from_date: str = None,
                         to_date: str = None,
                         employee_id: str = None,
                         office_id: str = None,
                         department_id: str = None) -> Dict:
        """Get data collection metrics without generating full report"""

        metrics = {
            'employees': {
                'total': 0,
                'filtered': 0,
                'pages': 0
            },
            'time_entries': {
                'total': 0,
                'pages': 0
            },
            'api_calls': 0,
            'estimated_time': 0
        }

        try:
            self.logger.info("=== OBTENIENDO MÉTRICAS DE DATOS ===")

            # Step 1: Get ALL employees with pagination
            self.logger.info("PASO 1: Contando empleados...")
            if employee_id:
                metrics['employees']['total'] = 1
                metrics['employees']['filtered'] = 1
                metrics['employees']['pages'] = 1
                self.api_calls += 1
            else:
                # Get ALL employees with pagination
                all_employees = []
                page = 1
                while True:
                    self.log_api_call(f"get_employees (page {page})")
                    employees_response = self.sesame_api.get_employees(
                        page=page, per_page=100)

                    if not employees_response or not employees_response.get(
                            'data'):
                        break

                    page_employees = employees_response['data']
                    all_employees.extend(page_employees)

                    meta = employees_response.get('meta', {})
                    total_pages = meta.get('totalPages', 1)
                    current_page = meta.get('page', 1)

                    self.logger.info(
                        f"  ✓ Página {current_page}/{total_pages}: {len(page_employees)} empleados"
                    )

                    if current_page >= total_pages:
                        break

                    page += 1

                metrics['employees']['total'] = len(all_employees)
                metrics['employees']['pages'] = page - 1

                # Apply filters
                employees = []
                for employee in all_employees:
                    include_employee = True

                    if office_id:
                        employee_office_id = employee.get(
                            'office',
                            {}).get('id') if employee.get('office') else None
                        if employee_office_id != office_id:
                            include_employee = False

                    if department_id and include_employee:
                        employee_department_id = employee.get(
                            'department', {}).get('id') if employee.get(
                                'department') else None
                        if employee_department_id != department_id:
                            include_employee = False

                    if include_employee:
                        employees.append(employee)

                metrics['employees']['filtered'] = len(employees)

                # Limit for testing
                if len(employees) > 10:
                    metrics['employees']['filtered'] = 10
                    employees = employees[:10]

            # Step 2: Sample a few employees to estimate time entries and breaks
            self.logger.info("PASO 2: Estimando registros de tiempo...")
            if len(employees) > 0:
                # Sample first 3 employees to estimate
                sample_employees = employees[:3]
                total_time_entries = 0
                total_time_pages = 0

                for employee in sample_employees:
                    emp_id = employee.get('id')
                    emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip(
                    )

                    # Get time entries for this employee
                    page = 1
                    employee_time_entries = 0
                    while True:
                        self.log_api_call(f"get_time_entries (page {page})",
                                          emp_name)
                        time_response = self.sesame_api.get_work_entries(
                            employee_id=emp_id,
                            from_date=from_date,
                            to_date=to_date,
                            page=page,
                            limit=300)

                        if not time_response or not time_response.get('data'):
                            break

                        page_data = time_response['data']
                        employee_time_entries += len(page_data)

                        meta = time_response.get('meta', {})
                        total_pages = meta.get('totalPages', 1)
                        current_page = meta.get('page', 1)

                        if current_page >= total_pages:
                            break

                        page += 1

                    total_time_entries += employee_time_entries
                    total_time_pages += (page - 1)

                # Estimate for all employees
                avg_time_entries = total_time_entries / len(sample_employees)
                avg_time_pages = total_time_pages / len(sample_employees)

                metrics['time_entries']['total'] = int(
                    avg_time_entries * metrics['employees']['filtered'])
                metrics['time_entries']['pages'] = int(
                    avg_time_pages * metrics['employees']['filtered'])

            metrics['api_calls'] = self.api_calls
            metrics['estimated_time'] = int(time.time() - self.start_time)

            self.logger.info(f"=== MÉTRICAS COMPLETADAS ===")
            self.logger.info(
                f"Empleados: {metrics['employees']['filtered']}/{metrics['employees']['total']}"
            )
            self.logger.info(
                f"Time entries estimados: {metrics['time_entries']['total']}")
            self.logger.info(f"API calls realizadas: {metrics['api_calls']}")

            return metrics

        except Exception as e:
            self.logger.error(f"Error getting data metrics: {str(e)}")
            return metrics

    def generate_optimized_report(
            self,
            from_date: str = None,
            to_date: str = None,
            employee_id: str = None,
            office_id: str = None,
            department_id: str = None,
            report_type: str = "by_employee") -> Optional[bytes]:
        """Generate optimized report with complete pagination"""

        try:
            self.logger.info(
                "=== INICIANDO GENERACIÓN DE REPORTE OPTIMIZADO ===")
            self.logger.info(
                f"Parámetros: from_date={from_date}, to_date={to_date}, employee_id={employee_id}"
            )

            # Step 1: Get ALL employees with pagination
            self.logger.info("PASO 1: Obteniendo TODOS los empleados...")
            if employee_id:
                self.log_api_call("get_employee_details", employee_id)
                employee_response = self.sesame_api.get_employee_details(
                    employee_id)
                employees = [employee_response.get('data')
                             ] if employee_response else []
            else:
                # Get ALL employees with pagination
                all_employees = []
                page = 1
                while True:
                    self.log_api_call(f"get_employees (page {page})")
                    employees_response = self.sesame_api.get_employees(
                        page=page, per_page=100)

                    if not employees_response or not employees_response.get(
                            'data'):
                        break

                    page_employees = employees_response['data']
                    all_employees.extend(page_employees)

                    # Check if there are more pages
                    meta = employees_response.get('meta', {})
                    total_pages = meta.get('totalPages', 1)
                    current_page = meta.get('page', 1)
                    total_records = meta.get('total', len(page_employees))

                    self.logger.info(
                        f"  ✓ Página {current_page}/{total_pages}: {len(page_employees)} empleados (Total acumulado: {len(all_employees)}/{total_records})"
                    )

                    if current_page >= total_pages:
                        break

                    page += 1

                self.logger.info(
                    f"  ✓ TOTAL: {len(all_employees)} empleados obtenidos")

                # Apply filters
                employees = []
                for employee in all_employees:
                    include_employee = True

                    if office_id:
                        employee_office_id = employee.get(
                            'office',
                            {}).get('id') if employee.get('office') else None
                        if employee_office_id != office_id:
                            include_employee = False

                    if department_id and include_employee:
                        employee_department_id = employee.get(
                            'department', {}).get('id') if employee.get(
                                'department') else None
                        if employee_department_id != department_id:
                            include_employee = False

                    if include_employee:
                        employees.append(employee)

                if len(employees) != len(all_employees):
                    self.logger.info(
                        f"  ✓ Después de filtros: {len(employees)} empleados seleccionados"
                    )

                # TEMPORAL: Limitar empleados para evitar timeout durante pruebas
                if len(employees) > 10:
                    self.logger.info(
                        f"  ⚠️ LIMITANDO a 10 empleados para evitar timeout (total disponible: {len(employees)})"
                    )
                    employees = employees[:10]

            if not employees:
                self.logger.warning("No employees found")
                return self._create_empty_report()

            self.logger.info(
                f"✓ Encontrados {len(employees)} empleados para procesar")

            # Step 2: Get check types once
            self.logger.info("PASO 2: Obteniendo tipos de actividad...")
            self.log_api_call("get_check_types")
            check_types_response = self.sesame_api.get_check_types(
                page=1, per_page=100)
            check_types_map = {}
            if check_types_response and check_types_response.get('data'):
                for check_type in check_types_response['data']:
                    check_types_map[check_type.get('id')] = check_type.get(
                        'name', 'Actividad no especificada')
                self.logger.info(
                    f"✓ Cargados {len(check_types_map)} tipos de actividad")

            # Step 3: Process each employee with complete pagination
            self.logger.info("PASO 3: Procesando empleados...")
            all_employee_data = {}

            for i, employee in enumerate(employees, 1):
                emp_id = employee.get('id')
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip(
                )

                self.logger.info(
                    f"--- Procesando empleado {i}/{len(employees)}: {emp_name} ---"
                )

                # Get ALL time entries with pagination
                time_entries = self._get_all_paginated_data(
                    endpoint_func=self.sesame_api.get_work_entries,
                    employee_id=emp_id,
                    from_date=from_date,
                    to_date=to_date,
                    data_type="time entries",
                    employee_name=emp_name)

                all_employee_data[emp_id] = {
                    'employee': employee,
                    'time_entries': time_entries
                }

            # Step 4: Create Excel report
            self.logger.info("PASO 4: Generando archivo Excel...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Optimizado"

            # Headers
            headers = [
                "Empleado", "Tipo ID", "Número ID", "Fecha", "Actividad",
                "Grupo", "Entrada", "Salida", "Tiempo Registrado"
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC",
                                        end_color="CCCCCC",
                                        fill_type="solid")

            current_row = 2
            total_records = 0

            # Process employee data
            for emp_id, data in all_employee_data.items():
                employee = data['employee']
                emp_name = f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip(
                )
                identity_type, identity_number = self._get_employee_identification(
                    employee)

                time_entries = data['time_entries']

                if time_entries:
                    # Group by date
                    entries_by_date = {}
                    for entry in time_entries:
                        work_in = entry.get(
                            'workEntryIn',
                            {}) if entry.get('workEntryIn') else {}
                        if work_in and work_in.get('date'):
                            try:
                                parsed_date = self._parse_datetime(
                                    work_in.get('date'))
                                if parsed_date:
                                    date_key = parsed_date.strftime('%Y-%m-%d')
                                    if date_key not in entries_by_date:
                                        entries_by_date[date_key] = []
                                    entries_by_date[date_key].append(entry)
                            except Exception:
                                continue

                    # Process each date
                    for date_key in sorted(entries_by_date.keys()):
                        date_entries = entries_by_date[date_key]

                        for entry in date_entries:
                            work_in = entry.get(
                                'workEntryIn',
                                {}) if entry.get('workEntryIn') else {}
                            work_out = entry.get(
                                'workEntryOut',
                                {}) if entry.get('workEntryOut') else {}

                            start_time = self._parse_datetime(
                                work_in.get('date')) if work_in else None
                            end_time = self._parse_datetime(
                                work_out.get('date')) if work_out else None

                            # Skip entries without start time
                            if not start_time:
                                continue

                            # Calculate duration
                            duration_str = "00:00:00"
                            end_time_str = "--"

                            if start_time and end_time:
                                duration = end_time - start_time
                                entry_seconds = duration.total_seconds()
                                hours = int(entry_seconds // 3600)
                                minutes = int((entry_seconds % 3600) // 60)
                                seconds = int(entry_seconds % 60)
                                duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                                end_time_str = end_time.strftime("%H:%M:%S")
                            elif start_time and not end_time:
                                # Si no hay hora de salida, mostrar guión
                                duration_str = "--"
                                end_time_str = "--"

                            # Get activity name
                            activity_name = entry.get('workEntryType',
                                                      'Trabajo')

                            # Add row
                            ws.cell(row=current_row, column=1, value=emp_name)
                            ws.cell(row=current_row,
                                    column=2,
                                    value=identity_type)
                            ws.cell(row=current_row,
                                    column=3,
                                    value=identity_number)
                            ws.cell(row=current_row, column=4, value=date_key)
                            ws.cell(row=current_row,
                                    column=5,
                                    value=activity_name)
                            ws.cell(row=current_row, column=6, value="")
                            ws.cell(row=current_row,
                                    column=7,
                                    value=start_time.strftime("%H:%M:%S"))
                            ws.cell(row=current_row,
                                    column=8,
                                    value=end_time_str)
                            ws.cell(row=current_row,
                                    column=9,
                                    value=duration_str)
                            current_row += 1
                            total_records += 1
                else:
                    # No time entries
                    ws.cell(row=current_row, column=1, value=emp_name)
                    ws.cell(row=current_row, column=2, value=identity_type)
                    ws.cell(row=current_row, column=3, value=identity_number)
                    ws.cell(row=current_row, column=4, value="Sin datos")
                    ws.cell(row=current_row,
                            column=5,
                            value="No hay registros")
                    ws.cell(row=current_row, column=6, value="")
                    ws.cell(row=current_row, column=7, value="--")
                    ws.cell(row=current_row, column=8, value="--")
                    ws.cell(row=current_row, column=9, value="00:00:00")
                    current_row += 1

            # Add summary row
            ws.cell(row=current_row, column=1, value="RESUMEN")
            ws.cell(row=current_row,
                    column=2,
                    value=f"Empleados: {len(employees)}")
            ws.cell(row=current_row,
                    column=3,
                    value=f"Registros: {total_records}")
            ws.cell(row=current_row,
                    column=4,
                    value=f"API calls: {self.api_calls}")
            ws.cell(row=current_row,
                    column=5,
                    value=f"Tiempo: {time.time() - self.start_time:.1f}s")

            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            total_time = time.time() - self.start_time
            self.logger.info(f"=== REPORTE OPTIMIZADO COMPLETADO ===")
            self.logger.info(f"Total API calls: {self.api_calls}")
            self.logger.info(f"Total time: {total_time:.1f}s")
            self.logger.info(f"Empleados procesados: {len(employees)}")
            self.logger.info(f"Registros generados: {total_records}")

            return output.getvalue()

        except Exception as e:
            self.logger.error(f"Error generating optimized report: {str(e)}")
            import traceback
            traceback.print_exc()
            return self._create_error_report(str(e))

    def _get_all_paginated_data(self, endpoint_func, employee_id, from_date,
                                to_date, data_type, employee_name):
        """Get all paginated data for a specific endpoint"""
        all_data = []
        page = 1

        while True:
            self.log_api_call(
                f"get_{data_type.replace(' ', '_')} (page {page})",
                employee_name)

            response = endpoint_func(employee_id=employee_id,
                                     from_date=from_date,
                                     to_date=to_date,
                                     page=page,
                                     limit=100)

            if not response or not response.get('data'):
                break

            page_data = response['data']
            all_data.extend(page_data)

            # Check if there are more pages
            meta = response.get('meta', {})
            total_pages = meta.get('totalPages', 1)
            current_page = meta.get('page', 1)
            total_records = meta.get('total', len(page_data))

            self.logger.info(
                f"  ✓ Página {current_page}/{total_pages}: {len(page_data)} {data_type} (Total acumulado: {len(all_data)}/{total_records})"
            )

            if current_page >= total_pages:
                break

            page += 1

        if all_data:
            self.logger.info(
                f"  ✓ TOTAL: {len(all_data)} {data_type} obtenidos")
        else:
            self.logger.info(f"  ✗ No {data_type}")

        return all_data

    def _create_empty_report(self) -> bytes:
        """Create an empty report when no data is found"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Vacío"

        ws.cell(row=1, column=1, value="No se encontraron empleados")
        ws.cell(row=2, column=1, value="Verifique los filtros aplicados")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_error_report(self, error_message: str) -> bytes:
        """Create an error report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Error"

        ws.cell(row=1, column=1, value="Error en generación optimizada")
        ws.cell(row=2, column=1, value=error_message)
        ws.cell(row=3,
                column=1,
                value=f"API calls realizadas: {self.api_calls}")
        ws.cell(
            row=4,
            column=1,
            value=f"Tiempo transcurrido: {time.time() - self.start_time:.1f}s")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_datetime(self, date_str: str) -> Optional[datetime]:
        """Parse datetime string from API"""
        if not date_str:
            return None

        try:
            return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        except ValueError:
            try:
                return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None

    def _get_employee_identification(self, employee: Dict) -> tuple:
        """Extract employee identification type and number"""
        identification_number = (
            employee.get('nid') or employee.get('identificationNumber')
            or employee.get('code') or str(employee.get('pin')) if
            employee.get('pin') else None or employee.get('id', 'Sin número'))

        identification_type = (employee.get('identityNumberType')
                               or employee.get('identificationType') or 'DNI')

        return identification_type, str(identification_number)
